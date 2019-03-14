from openpyxl import Workbook, load_workbook
import pandas as pd

file = pd.read_csv('C.csv')
b = file['Variety Name']

print(b)

print(b.shape[0])
length = b.shape[0]

wb = load_workbook('C3.xlsx')
ws = wb['Sheet1']
for i in range(2,length+1):
    if type(ws.cell(row = i, column= 8).value) == str:
        temp = ws.cell(row = i, column= 8).value.lower()
        ws.cell(row=i, column=8, value=temp)
    print(ws.cell(row=i, column=8).value)
wb.save('C3.xlsx')