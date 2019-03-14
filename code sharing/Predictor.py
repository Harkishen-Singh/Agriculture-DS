from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd

class Predictor():
    def __init__(self):
        referal_book = input('Enter the referal book name (without any extension) : ')

        self.ratio_bad3 = load_workbook('ratios/ratio_bad_'+referal_book+'.xlsx')
        self.ratio_normal3 = load_workbook('ratios/ratio_normal_' + referal_book + '.xlsx')
        self.ratio_good3 = load_workbook('ratios/ratio_good_' + referal_book + '.xlsx')
        self.ratio_good2 = self.ratio_good3['Sheet']
        self.ratio_normal2 = self.ratio_normal3['Sheet']
        self.ratio_bad2 = self.ratio_bad3['Sheet']
        self.ratio_bad = pd.read_excel('ratios/ratio_bad_'+referal_book+'.xlsx')
        self.ratio_normal = pd.read_excel('ratios/ratio_normal_' + referal_book + '.xlsx')
        self.ratio_good = pd.read_excel('ratios/ratio_good_' + referal_book + '.xlsx')


        self.blockname = ''
        self.vareityname = ''
        self.systemCultivation = ''
        self.isIrrigated = ''
        self.waterSource = ''
        self.pestDamage = ''
        self.operationSize = 0  # in hectare
        self.cultivatedSize = 0  # in hectare
        self.EnterValues()

    def EnterValues(self):
        print('Please input the required values')

        self.blockname = input('Block Name : ')
        #self.vareityname = input('Variety Name : ').lower
        self.vareityname = 'puja'
        #self.systemCultivation = input('System of Cultivation : ').lower
        self.systemCultivation = 'conventional'
        #self.isIrrigated = input('is irrigated : ').lower
        self.isIrrigated = 'rainfed'
        #self.yieldingType = input('yielding type').lower()
        self.yieldingType = 'local'
        #self.waterSource = input('Water Source : ')
        self.pestDamage = int(input('Pest Damage in Integers : '))
        self.operationSize = float(input('Operational Size holding (in hectares) : '  ))
        self.cultivatedSize = float(input('Operational Size in Cultivation (in hectares) : '  ))
        self.calculatingWeightage()

    def calculatingWeightage(self):

        self.varietyRow_number = [0]*3
        i =1
        check = False
        while True:
            i += 1
            if self.ratio_bad2.cell(row=i, column=4).value == None:
                break

            v = self.ratio_bad2.cell(row=i, column=4).value
            if self.vareityname == v:
                self.varietyRow_number[0] = i
                check = True
                break
        if check == False:
            print('Variety not found in the bads file')
        i = 1
        check = False
        while True:
            i += 1
            if self.ratio_normal2.cell(row=i, column=4).value == None:
                break

            v = self.ratio_normal2.cell(row = i, column = 4).value
            if self.vareityname == v:
                self.varietyRow_number[1] = i
                check = True
                break
        if check == False:
            print('Variety not found in the normals file')
        i = 1
        check = False
        while True:
            i += 1
            if self.ratio_good2.cell(row=i, column=4).value == None:
                break

            v = self.ratio_good2.cell(row = i, column = 4).value
            if self.vareityname == v:
                self.varietyRow_number[2] = i
                check = True
                break
        if check == False:
            print('Variety not found in the goods file')

        i = 1

        ### *********** deciding Category ***************

        # bad file part
        sum1 =0
        if self.isIrrigated == 'irrigated':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=8).value
        elif self.isIrrigated == 'rainfed':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=9).value
        elif self.isIrrigated == 'unirrigated':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=10).value
        else:
            print('Wrong input in irrigation part')
            exit(0)

        if self.systemCultivation == 'conventional':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=11).value
        elif self.systemCultivation == 'sri':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=12).value
        else:
            print('wrong input in method of cultivation part')
            exit(0)

        if self.yieldingType == 'high yielding':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=13).value
        elif self.yieldingType == 'local':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=14).value
        elif self.yieldingType == 'hybrid':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=15).value
        else:
            print('wrong input in variety part')
            exit(0)

        # normal file part

        sum2 =0
        if self.isIrrigated == 'irrigated':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=8).value
        elif self.isIrrigated == 'rainfed':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=9).value
        elif self.isIrrigated == 'unirrigated':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=10).value
        else:
            print('Wrong input in irrigation part')
            exit(0)

        if self.systemCultivation == 'conventional':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=11).value
        elif self.systemCultivation == 'sri':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=12).value
        else:
            print('wrong input in method of cultivation part')
            exit(0)

        if self.yieldingType == 'high yielding':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=13).value
        elif self.yieldingType == 'local':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=14).value
        elif self.yieldingType == 'hybrid':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=15).value
        else:
            print('wrong input in variety part')
            exit(0)

        # good file part

        sum3 = 0
        if self.isIrrigated == 'irrigated':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=8).value
        elif self.isIrrigated == 'rainfed':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=9).value
        elif self.isIrrigated == 'unirrigated':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=10).value
        else:
            print('Wrong input in irrigation part')
            exit(0)

        if self.systemCultivation == 'conventional':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=11).value
        elif self.systemCultivation == 'sri':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=12).value
        else:
            print('wrong input in method of cultivation part')
            exit(0)

        if self.yieldingType == 'high yielding':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=13).value
        elif self.yieldingType == 'local':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=14).value
        elif self.yieldingType == 'hybrid':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=15).value
        else:
            print('wrong input in variety part')
            exit(0)

        ### *********** pest importance calculation ****************

        val = self.pestDamage
        valpest_bad = self.ratio_bad2.cell(row=self.varietyRow_number[0], column=16).value
        valpest_normal = self.ratio_normal2.cell(row=self.varietyRow_number[1], column=16).value
        valpest_good = self.ratio_good2.cell(row=self.varietyRow_number[2], column=16).value

        d_bad = abs(val-valpest_bad)
        d_normal = abs(val-valpest_normal)
        d_good = abs(val-valpest_good)
        result = ''

        if d_bad < d_normal and d_bad < d_good:
            sum1 += 2
            result = 'bad'
        elif d_normal < d_bad and d_normal < d_good :
            sum2 += 2
            result = 'normal'
        elif d_good < d_bad and d_good < d_normal:
            sum3 += 2
            result = 'good'
        else:
            print('Cannot decide the pestDamage Factor. Please Report the Bug team.')



        print('Sum 1 :'+str(sum1) + '\nSum 2 : ' + str(sum2) + '\nSum 3 : '+str(sum3))

        if result == 'bad':
            self.predictingBad(self.varietyRow_number[0])
        elif result == 'normal':
            self.predictingNormal(self.varietyRow_number[1])
        elif result == 'good':
            self.predictingGood(self.varietyRow_number[2])
        else:
            print('Result Calculation Failed.! Please report the Bug Team about this.')

        
    def predictingBad(self, rowNumber):
        pre_greenWtProduced = self.ratio_bad2.cell(row = rowNumber, column=1).value * self.cultivatedSize
        pre_dryWt = self.ratio_bad2.cell(row = rowNumber, column=2).value * self.cultivatedSize
        pre_normalYieldinKilo = self.ratio_bad2.cell(row = rowNumber, column=3).value * self.cultivatedSize

        print('Expected Values :')
        print('Green Weight Produced would be ' + str(pre_greenWtProduced)+'\nDry Weight Produced would be '+str(pre_dryWt)\
              +'\nNormal Yield in Kilograms would be '+str(pre_normalYieldinKilo)
              )

    def predictingNormal(self, rowNumber):
        pre_greenWtProduced = self.ratio_normal2.cell(row=rowNumber, column=1).value * self.cultivatedSize
        pre_dryWt = self.ratio_normal2.cell(row=rowNumber, column=2).value * self.cultivatedSize
        pre_normalYieldinKilo = self.ratio_normal2.cell(row=rowNumber, column=3).value * self.cultivatedSize

        print('Expected Values :')
        print('Green Weight Produced would be ' + str(pre_greenWtProduced) + '\nDry Weight Produced would be ' + str(pre_dryWt) \
              + '\nNormal Yield in Kilograms would be ' + str(pre_normalYieldinKilo)
              )

    def predictingGood(self, rowNumber):
        pre_greenWtProduced = self.ratio_good2.cell(row=rowNumber, column=1).value * self.cultivatedSize
        pre_dryWt = self.ratio_good2.cell(row=rowNumber, column=2).value * self.cultivatedSize
        pre_normalYieldinKilo = self.ratio_good2.cell(row=rowNumber, column=3).value * self.cultivatedSize

        print('Expected Values :')
        print('Green Weight Produced would be ' + str(pre_greenWtProduced) + '\nDry Weight Produced would be ' + str(pre_dryWt) \
              + '\nNormal Yield in Kilograms would be ' + str(pre_normalYieldinKilo)
              )






if __name__ == '__main__':
    obj = Predictor()