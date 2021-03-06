from Sorter import Sorter
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook

class Process(Sorter):

    def xx(self):
        '''This class is the main backboe of ML from the excel database inputs'''
        self.filename = self.sortedFileName
        self.wb3 = load_workbook(self.sortedFileName)
        self.ws3 = self.wb3['Sorted']
        self.row = 2

        self.file = Workbook()
        self.sheet = self.file.active
        self.sheet.cell(row = 1, column = 1, value='greenWtProduced_ratio')
        self.sheet.cell(row=1, column=2, value='DryWtProduced_ratio')
        self.sheet.cell(row=1, column=3, value='normalYldKilo_ratio')
        self.sheet.cell(row=1, column=4, value='Variety_Name')
        while(self.row):
            if self.ws3.cell(row = self.row, column = 2).value != None:
                self.row = self.row + 1
            else:
                break
        #print('Number of rows in sorted = '+ str(self.row))
        self.processing()

    def processing(self):
        sg = 0 ;sd = 0; sn = 0; counter = 2
        ag = 0; ad = 0;an = 0
        checker = True
        for i in range(2, self.row):
            no = i
            if checker == True:
                if type(self.ws3.cell(row=no, column=10).value) == int:

                    seedExtractor = self.ws3.cell(row=no, column=5).value
                    l2 = len(seedExtractor)
                    number = ''
                    for uu in seedExtractor:
                        if uu.isdigit() or uu == '.' :
                            number = number + uu
                    p1 = float(number)
                    number = ''

                    seedExtractor2 = self.ws3.cell(row=no, column=10).value
                    if type(seedExtractor2) != int:

                        for uu2 in seedExtractor2:
                            if uu2.isdigit() or uu2 == '.'  : number = number + uu2
                        p2 = float(number)
                        number = ''
                        #print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                    if type(seedExtractor2) == int:
                        p2 = seedExtractor2
                        #print('p1 is '+str(p1) + ' p2(int wala) is '+str(p2))

                    self.totalseeds = p1 * p2
                else:
                    if type(self.ws3.cell(row=no, column=10).value) == int:

                        seedExtractor = self.ws3.cell(row=no, column=5).value
                        l2 = len(seedExtractor)
                        number = ''
                        for uu in seedExtractor:
                            if uu.isdigit() or uu == '.' :
                                number = number + uu
                        p1 = float(number)
                        number = ''

                        seedExtractor2 = self.ws3.cell(row=no, column=10).value
                        if type(seedExtractor2) != int:

                            for uu2 in seedExtractor2:
                                if uu2.isdigit() or uu2 == '.' : number = number + uu2
                            p2 = float(number)
                            number = ''
                            #print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                        if type(seedExtractor2) == int:
                            p2 = seedExtractor2
                            #print('p1 is ' + str(p1) + ' p2(int wala) is ' + str(p2))

                        self.totalseeds = p1 * p2
                #print(str(self.ws3.cell(row=no, column=15).value) + ' ' + str(self.totalseeds))
                #print(str(self.ws3.cell(row=no, column=15).value) + ' here is it' + str(i))
                self.greenWtProduced_ratio = self.ws3.cell(row=no, column=15).value / self.totalseeds
                self.DryWtProduced_ratio = self.ws3.cell(row=no, column=16).value / self.totalseeds
                self.normalYldKilo_ratio = self.ws3.cell(row=no, column=18).value / self.totalseeds
                sg = sg + self.greenWtProduced_ratio;
                ag = ag + 1
                sd = sd + self.DryWtProduced_ratio;
                ad = ad + 1
                sn = sn + self.normalYldKilo_ratio;
                an = an + 1
                checker = False

            elif self.ws3.cell(row = no, column = 8).value == self.ws3.cell(row = no - 1, column = 8).value :
                if type(self.ws3.cell(row=no, column=10).value) == int:

                    seedExtractor = self.ws3.cell(row=no, column=5).value
                    l2 = len(seedExtractor)
                    number = ''
                    for uu in seedExtractor:
                        if uu.isdigit() or uu == '.' :
                            number = number + uu
                    p1 = float(number)
                    number = ''

                    seedExtractor2 = self.ws3.cell(row=no, column=10).value
                    if type(seedExtractor2) != int:

                        for uu2 in seedExtractor2:
                            if uu2.isdigit() or uu2 == '.'  : number = number + uu2
                        p2 = float(number)
                        number = ''
                        #print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                    if type(seedExtractor2) == int:
                        p2 = seedExtractor2
                        #print('p1 is '+str(p1) + ' p2(int wala) is '+str(p2))

                    self.totalseeds = p1 * p2
                else:
                    if type(self.ws3.cell(row=no, column=10).value) == int:

                        seedExtractor = self.ws3.cell(row=no, column=5).value
                        l2 = len(seedExtractor)
                        number = ''
                        for uu in seedExtractor:
                            if uu.isdigit() or uu == '.' :
                                number = number + uu
                        p1 = float(number)
                        number = ''

                        seedExtractor2 = self.ws3.cell(row=no, column=10).value
                        if type(seedExtractor2) != int:

                            for uu2 in seedExtractor2:
                                if uu2.isdigit() or uu2 == '.' : number = number + uu2
                            p2 = float(number)
                            number = ''
                            #print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                        if type(seedExtractor2) == int:
                            p2 = seedExtractor2
                            #print('p1 is ' + str(p1) + ' p2(int wala) is ' + str(p2))

                        self.totalseeds = p1 * p2
                #print(str(self.ws3.cell(row=no, column=15).value) + ' ' + str(self.totalseeds))

                self.greenWtProduced_ratio =  self.ws3.cell(row = no, column = 15).value / self.totalseeds
                self.DryWtProduced_ratio = self.ws3.cell(row = no, column = 16).value / self.totalseeds
                self.normalYldKilo_ratio = self.ws3.cell(row = no, column = 18).value / self.totalseeds
                sg = sg + self.greenWtProduced_ratio; ag = ag +1
                sd = sd + self.DryWtProduced_ratio; ad = ad +1
                sn = sn + self.normalYldKilo_ratio; an = an +1

            if self.ws3.cell(row = no, column = 8).value != self.ws3.cell(row = no + 1, column = 8).value :
                self.sheet.cell(row=counter, column=1, value=sg / ag); ag = 0; sg = 0
                self.sheet.cell(row=counter, column=2, value=sd / ad); ad = 0; sd = 0
                self.sheet.cell(row=counter, column=3, value=sn / an); an = 0; sn = 0
                self.sheet.cell(row=counter, column=4, value=self.ws3.cell(row = no, column = 8).value)
                checker = True
                counter = counter + 1



        self.saver()

    def saver(self):
        #print('Number of varities : '+ str(len(self.varietyNameOnly)))
        #print(self.varietyNameOnly)
        self.file.save('ratio_saver2.xlsx')






obj = Process()
obj.individualAllotment()
obj.process()
obj.xx()
