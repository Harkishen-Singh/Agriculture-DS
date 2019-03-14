from openpyxl import load_workbook
import math

name = input('Enter the workbook name without extension(should be of .xlsx type) : ')
name = name + '.xlsx'
wb = load_workbook(name)
ws = wb['Sheet1']
i = 1
number = ''
num = 0.0
xxx = 2
cc= 0
ss = 0
while True:

    i += 1
    if ws.cell(row = i, column = 5).value == None or ws.cell(row = i, column = 15).value == None:
        break
    checker1 = False
    print(str(i) + ' ' +ws.cell(row = i, column = 5).value + ' ' + str(type(ws.cell(row = i, column = 5).value)))
    for x in ws.cell(row = i, column = 5).value :

        if  x.isdigit() or x == '.':
            number += x
            #print(number + ' break by break')

        elif x.isalpha() and x != ' ' and x != '§':
            cc +=1
            if x == 'A' or x == 'S' and checker1 == False and cc == 1:
                print(x)
                checker1 = True
                print(number + ' this ')
                num = float(number)
                num = num * 0.40468564
                num = round(num,5)
                ws.cell(row=i, column=5, value=str(num) + ' Hectares')
                print(num)
                num = 0
                number = ''
                cc= 0
                break
            else:
                number = ''



    checker2 = False
    for x in ws.cell(row=i, column=15).value:

        if x.isdigit() or x.isdecimal():
            number += x

        elif x.isalpha() and x != ' ' and x != '§':
            ss += 1
            if x == 'A' or x == 'S' and checker2 == False and ss == 1:
                checker2 = True
                num = float(number)
                num = num * 0.40468564
                num = round(num, 5)
                ws.cell(row=i, column=15, value=str(num) + ' Hectares')
                num = 0
                number = ''
                ss = 0
                break
            else:
                number = ''



wb.save('optimised_'+ name)
