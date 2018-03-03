from openpyxl import Workbook, load_workbook
from Base import Base

class Sorter(Base):
    def process(self):
        self.sheetname = self.ws
        wb2 = Workbook()
        self.ws2 = wb2.active

        '''for i in range(0, len(self.headingIndex)): # alloting headings to the new spreadsheet file
            self.ws2.cell(row = 1, column= i+1, value = self.headingsSheet[i])'''
        newRowCount = 2
        a = len(self.varietyNameOnly)
        b = a
        for x in range(0, len(self.varietyNameOnly)):

            for i in range(2, self.numberOfRows):

                if self.sheetname.cell(row=i, column=self.headingIndex[6]).value == self.varietyNameOnly[x]:
                    #print('matched with ' +str(self.sheetname.cell(row=i, column=self.headingIndex[6]).value) + ' '+str(self.varietyNameOnly[x]))

                    #print('i is : '+ str(i))
                    self.ws2.cell(row = newRowCount, column= 1, value= None)

                    self.ws2.cell(row=newRowCount, column=2, value=self.blockName[i-2])
                    self.ws2.cell(row=newRowCount, column=3, value=self.villageName[i-2])
                    self.ws2.cell(row=newRowCount, column=4, value=self.areaHolding[i-2])
                    self.ws2.cell(row=newRowCount, column=5, value=self.areaCultivation[i-2])
                    self.ws2.cell(row=newRowCount, column=6, value=self.sysCultivation[i-2])
                    self.ws2.cell(row=newRowCount, column=7, value=self.cropVariety[i-2])
                    self.ws2.cell(row=newRowCount, column=8, value=self.varietyName[i-2])
                    self.ws2.cell(row=newRowCount, column=9, value=self.sourceSeed[i-2])
                    self.ws2.cell(row=newRowCount, column=10, value=self.seedpHectare[i-2])
                    self.ws2.cell(row=newRowCount, column=11, value=self.manureFYMq[i-2])
                    self.ws2.cell(row=newRowCount, column=12, value=self.chemFertiQ[i-2])
                    self.ws2.cell(row=newRowCount, column=13, value=self.dateTransplanting[i-2])
                    self.ws2.cell(row=newRowCount, column=14, value=self.dateHarvest[i-2])
                    self.ws2.cell(row=newRowCount, column=15, value=self.greenWtCCE[i-2])
                    self.ws2.cell(row=newRowCount, column=16, value=self.dryWtCCE[i-2])
                    self.ws2.cell(row=newRowCount, column=17, value=self.moisturePercent[i-2])
                    self.ws2.cell(row=newRowCount, column=18, value=self.normalYldKilo[i-2])
                    self.ws2.cell(row=newRowCount, column=19, value=self.proQuantityUser[i-2])
                    self.ws2.cell(row=newRowCount, column=20, value=self.proQualityUser[i-2])
                    self.ws2.cell(row=newRowCount, column=21, value=self.isIrrigated[i-2])
                    self.ws2.cell(row=newRowCount, column=22, value=self.waterSource[i-2])
                    self.ws2.cell(row=newRowCount, column=23, value=self.weatherCond[i-2])
                    self.ws2.cell(row=newRowCount, column=24, value=self.pest[i-2])
                    self.ws2.cell(row=newRowCount, column=25, value=self.stress[i-2])
                    self.ws2.cell(row=newRowCount, column=25, value=self.dateCutting[i-2])
                    newRowCount = newRowCount + 1
                    #print(self.blockName[i-2] + ' '+ self.villageName[i-2]+" "+self.areaHolding[-2]+' '+str(self.varietyName[i-2]))

        #print(self.headingsSheet)
        self.ws2.title = 'Sorted'
        wb2.save('sortedFile'+self.workbook_name[2:])
        print('Sorted Data saved at '+ 'sortedFile'+self.workbook_name[2:])
        self.sortedFileName = 'sortedFile'+self.workbook_name[2:]

