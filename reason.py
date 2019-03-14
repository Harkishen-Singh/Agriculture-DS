from manures import Manures
from openpyxl import load_workbook, Workbook
import csv

class Reason(Manures):
    def format_check(self):
        file1 = load_workbook(self.sortedFileName)
        print(self.sortedFileName)
        sheet1 = file1['Sorted']
        count = 2
        isWrong = False
        while sheet1.cell(row=count, column=2).value != None :

            cross = sheet1.cell(row=count, column=21).value
            if cross == None: break
            #print(count)
            if cross.lower() != 'irrigated' and cross.lower() != 'rainfed' and cross.lower() != 'un-irrigated' \
                    and cross.lower() != 'unirrigated' :
                isWrong = True
                break
            count = count + 1
        if isWrong == True:
            print('The file '+self.sortedFileName + ' doesnot contain the required data in column 21 in the required format.')
            print('Irrigation format wrong')
            exit(0)
        isWrong = False
        count = 2
        while sheet1.cell(row=count, column=6).value != None:
            count = count + 1
            cross = sheet1.cell(row=count, column=6).value
            if cross == None: break
            if cross.lower() != 'conventional' and cross.lower() != 'sri' :
                isWrong = True
                break
            count = count + 1
        if isWrong == True:
            print('The file ' + self.sortedFileName + ' doesnot contain the required data in column 21 in the required format.')
            print('seeds detail format first part wrong')
            exit(0)
        isWrong = False
        count = 2
        while sheet1.cell(row=count, column=7).value != None:
            count = count + 1
            cross = sheet1.cell(row=count, column=7).value
            if cross == None: break
            if cross.lower() != 'high yieldling' and cross.lower() != 'local' :
                isWrong = True
                break
            count = count + 1
        isWrong = False
        if isWrong == True :
            print('The file ' + self.sortedFileName + ' doesnot contain the required data in column 21 in the required format.')
            print('seeds details format second part wrong')
            exit(0)
        isWrong = False
        self.bad()


    def bad(self):
        file = load_workbook(self.sortedFileName_Bad)
        sheet = file['Sorted']
        self.isIrrigatedType = [0] * 3
        self.isIrrigatedType_counter=0

        # irrigated, rainfed, unirrigated
        self.seed_Details_part1 = [0] * 2
        self.seed_Details_part1_counter = 0

        # conventional , sri
        self.seed_Details_part2 = [0] * 3
        self.seed_Details_part2_counter = 0

        # high yielding, local, hybrid
        count = 1

        # pests damage avergae calculation
        self.pestDamage = 0
        self.pestDamage_counter = 0

        file2 = load_workbook('./ratios/ratio_bad_'+self.workbook_name2+'.xlsx')
        sheet2 = file2['Sheet']
        count2 = 1
        sheet2.cell(row=1, column=8, value='irrigated ratio')
        sheet2.cell(row=1, column=9, value='rainfed ratio')
        sheet2.cell(row=1, column=10, value='unirrigated ratio')
        sheet2.cell(row=1, column=11, value='conventional ratio')
        sheet2.cell(row=1, column=12, value='sri ratio')
        sheet2.cell(row=1, column=13, value='high yielding ratio')
        sheet2.cell(row=1, column=14, value='local ratio')
        sheet2.cell(row=1, column=15, value='hybrid ratio')
        sheet2.cell(row=1, column=16, value='pest damage')

        while sheet2.cell(row= count2, column=4).value != None:
            #print('Entered sheet2 part')
            count2 += 1
            variety = sheet2.cell(row= count2, column=4).value
            if variety == None: break

            checker2 = False
            count = 2
            while sheet.cell(row=count, column= 2).value != None:
                count = count + 1
                #print('sheet1 entered')
                # irrigation type part below
                if variety == sheet.cell(row = count, column=8 ).value :

                    checker2 = True
                    if sheet.cell(row = count, column= 21).value.lower() == 'irrigated':
                        self.isIrrigatedType[0] += 1
                        self.isIrrigatedType_counter += 1
                    elif sheet.cell(row=count, column= 21).value.lower() == 'rainfed':
                        self.isIrrigatedType[1] += 1
                        self.isIrrigatedType_counter +=1
                    elif sheet.cell(row=count, column= 21).value.lower() == 'un-irrigated'or \
                            sheet.cell(row=count, column=21).value.lower() == 'un irrigated':
                        self.isIrrigatedType[2] +=1
                        self.isIrrigatedType_counter += 1
                    else:
                        print('Some problem in checking the irrigation type from the database')
                        print('this : '+ sheet.cell(row=count, column= 21).value)

                    # seed details part1 below

                    if sheet.cell(row= count, column= 6).value.lower() == 'conventional':
                        self.seed_Details_part1[0] +=1
                        self.seed_Details_part1_counter +=1
                    elif sheet.cell(row= count, column= 6).value.lower() == 'sri':
                        self.seed_Details_part1[1] += 1
                        self.seed_Details_part1_counter += 1
                    else:
                        print('Some problem in checking the seed type part 1')

                    # seed details part2 below

                    if sheet.cell(row=count, column=7).value.lower() == 'high yielding variety':
                        self.seed_Details_part2[0] += 1
                        self.seed_Details_part2_counter += 1
                    elif sheet.cell(row=count, column=7).value.lower() == 'local':
                        self.seed_Details_part2[1] += 1
                        self.seed_Details_part2_counter += 1
                    elif sheet.cell(row=count, column=7).value.lower() == 'hybrid':
                        self.seed_Details_part2[2] += 1
                        self.seed_Details_part2_counter += 1

                    else:
                        print('Some problem in checking the seed type part 2')

                    if type(sheet.cell(row=count, column= 24).value) == float or type(sheet.cell(row=count, column=24).value) == int:
                        self.pestDamage += sheet.cell(row=count, column= 24).value
                        self.pestDamage_counter += 1


                if variety != sheet.cell(row = count, column=8 ).value and checker2==True:
                    #print('reached here 11')

                    ratio = self.isIrrigatedType[0] / self.isIrrigatedType_counter
                    sheet2.cell(row= count2, column=8, value=ratio)
                    #print(ratio)
                    ratio = self.isIrrigatedType[1] / self.isIrrigatedType_counter
                    sheet2.cell(row=count2, column=9, value=ratio)
                    #print(ratio)
                    ratio = self.isIrrigatedType[2] / self.isIrrigatedType_counter
                    sheet2.cell(row=count2, column=10, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part1[0] / self.seed_Details_part1_counter
                    sheet2.cell(row = count2, column= 11, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part1[1] / self.seed_Details_part1_counter
                    sheet2.cell(row=count2, column=12, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part2[0] / self.seed_Details_part2_counter
                    sheet2.cell(row=count2, column=13, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part2[1] / self.seed_Details_part2_counter
                    sheet2.cell(row=count2, column=14, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part2[2] / self.seed_Details_part2_counter
                    sheet2.cell(row=count2, column=15, value=ratio)
                    try :
                        ratio = self.pestDamage / self.pestDamage_counter
                    except ZeroDivisionError :
                        sheet2.cell(row=count2, column=16, value=0)
                    else:
                        sheet2.cell(row=count2, column=16, value=ratio)


                    ratio = 0
                    # initialising all to zero

                    self.isIrrigatedType = [0,0,0]; self.isIrrigatedType_counter = 0
                    self.seed_Details_part1 = [0,0]; self.seed_Details_part1_counter = 0
                    self.seed_Details_part2 = [0,0,0]; self.seed_Details_part2_counter = 0
                    self.pestDamage_counter = 0; self.pestDamage = 0
                    checker2 = False
                    break
                if sheet.cell(row=count, column=8).value == None: break
        file2.save('./ratios/ratio_bad_'+self.workbook_name2+'.xlsx')
        self.normal()

    def normal(self):
        file = load_workbook(self.sortedFileName_Normal)
        sheet = file['Sorted']
        self.isIrrigatedType = [0] * 3
        self.isIrrigatedType_counter=0

        # irrigated, rainfed, unirrigated
        self.seed_Details_part1 = [0] * 2
        self.seed_Details_part1_counter = 0

        # conventional , sri
        self.seed_Details_part2 = [0] * 3
        self.seed_Details_part2_counter = 0

        # high yielding, local, hybrid
        count = 1
        # pests damage avergae calculation
        self.pestDamage = 0
        self.pestDamage_counter = 0

        file2 = load_workbook('./ratios/ratio_normal_'+self.workbook_name2+'.xlsx')
        sheet2 = file2['Sheet']
        count2 = 1
        sheet2.cell(row=1, column=8, value='irrigated ratio')
        sheet2.cell(row=1, column=9, value='rainfed ratio')
        sheet2.cell(row=1, column=10, value='unirrigated ratio')
        sheet2.cell(row=1, column=11, value='conventional ratio')
        sheet2.cell(row=1, column=12, value='sri ratio')
        sheet2.cell(row=1, column=13, value='high yielding ratio')
        sheet2.cell(row=1, column=14, value='local ratio')
        sheet2.cell(row=1, column=15, value='hybrid ratio')
        sheet2.cell(row=1, column=16, value='pest damage')


        while sheet2.cell(row= count2, column=4).value != None:
            #print('Entered sheet2 part')
            count2 += 1
            variety = sheet2.cell(row= count2, column=4).value
            if variety == None: break

            checker2 = False
            count = 2
            while sheet.cell(row=count, column= 2).value != None:
                count = count + 1
                #print('sheet1 entered')
                # irrigation type part below
                if variety == sheet.cell(row = count, column=8 ).value :

                    checker2 = True
                    if sheet.cell(row = count, column= 21).value.lower() == 'irrigated':
                        self.isIrrigatedType[0] += 1
                        self.isIrrigatedType_counter += 1
                    elif sheet.cell(row=count, column= 21).value.lower() == 'rainfed':
                        self.isIrrigatedType[1] += 1
                        self.isIrrigatedType_counter +=1
                    elif sheet.cell(row=count, column= 21).value.lower() == 'un-irrigated'or \
                            sheet.cell(row=count, column=21).value.lower() == 'un irrigated':
                        self.isIrrigatedType[2] +=1
                        self.isIrrigatedType_counter += 1
                    else:
                        print('Some problem in checking the irrigation type from the database')
                        print('this : ' + sheet.cell(row=count, column=21).value)

                    # seed details part1 below

                    if sheet.cell(row= count, column= 6).value.lower() == 'conventional':
                        self.seed_Details_part1[0] +=1
                        self.seed_Details_part1_counter +=1
                    elif sheet.cell(row= count, column= 6).value.lower() == 'sri':
                        self.seed_Details_part1[1] += 1
                        self.seed_Details_part1_counter += 1
                    else:
                        print('Some problem in checking the seed type part 1')

                    # seed details part2 below

                    if sheet.cell(row=count, column=7).value.lower() == 'high yielding variety':
                        self.seed_Details_part2[0] += 1
                        self.seed_Details_part2_counter += 1
                    elif sheet.cell(row=count, column=7).value.lower() == 'local':
                        self.seed_Details_part2[1] += 1
                        self.seed_Details_part2_counter += 1
                    elif sheet.cell(row=count, column=7).value.lower() == 'hybrid':
                        self.seed_Details_part2[2] += 1
                        self.seed_Details_part2_counter += 1

                    else:
                        print('Some problem in checking the seed type part 2')

                    if type(sheet.cell(row=count, column= 24).value) == float or type(sheet.cell(row=count, column=24).value) == int:
                        self.pestDamage += sheet.cell(row=count, column= 24).value
                        self.pestDamage_counter += 1

                    ## also add an extension for hybrid

                if variety != sheet.cell(row = count, column=8 ).value and checker2==True:
                    #print('reached here 11')

                    ratio = self.isIrrigatedType[0] / self.isIrrigatedType_counter
                    sheet2.cell(row= count2, column=8, value=ratio)
                    #print(ratio)
                    ratio = self.isIrrigatedType[1] / self.isIrrigatedType_counter
                    sheet2.cell(row=count2, column=9, value=ratio)
                    #print(ratio)
                    ratio = self.isIrrigatedType[2] / self.isIrrigatedType_counter
                    sheet2.cell(row=count2, column=10, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part1[0] / self.seed_Details_part1_counter
                    sheet2.cell(row = count2, column= 11, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part1[1] / self.seed_Details_part1_counter
                    sheet2.cell(row=count2, column=12, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part2[0] / self.seed_Details_part2_counter
                    sheet2.cell(row=count2, column=13, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part2[1] / self.seed_Details_part2_counter
                    sheet2.cell(row=count2, column=14, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part2[2] / self.seed_Details_part2_counter
                    sheet2.cell(row=count2, column=15, value=ratio)

                    try :
                        ratio = self.pestDamage / self.pestDamage_counter
                    except ZeroDivisionError :
                        sheet2.cell(row=count2, column=16, value=0)
                    else:
                        sheet2.cell(row=count2, column=16, value=ratio)

                    ratio = 0


                    # initialising all to zero

                    self.isIrrigatedType = [0,0,0]; self.isIrrigatedType_counter = 0
                    self.seed_Details_part1 = [0,0]; self.seed_Details_part1_counter = 0
                    self.seed_Details_part2 = [0,0,0]; self.seed_Details_part2_counter = 0
                    self.pestDamage_counter = 0;
                    self.pestDamage = 0

                    checker2 = False
                    break
                if sheet.cell(row=count, column=8).value == None: break
        file2.save('./ratios/ratio_normal_'+self.workbook_name2+'.xlsx')
        self.good()

    def good(self):
        file = load_workbook(self.sortedFileName_Good)
        sheet = file['Sorted']
        self.isIrrigatedType = [0] * 3
        self.isIrrigatedType_counter=0

        # irrigated, rainfed, unirrigated
        self.seed_Details_part1 = [0] * 2
        self.seed_Details_part1_counter = 0

        # conventional , sri
        self.seed_Details_part2 = [0] * 3
        self.seed_Details_part2_counter = 0

        # high yielding, local, hybrid
        count = 1
        # pests damage avergae calculation
        self.pestDamage = 0
        self.pestDamage_counter = 0

        file2 = load_workbook('./ratios/ratio_good_'+self.workbook_name2+'.xlsx')
        sheet2 = file2['Sheet']
        count2 = 1
        sheet2.cell(row=1, column=8, value='irrigated ratio')
        sheet2.cell(row=1, column=9, value='rainfed ratio')
        sheet2.cell(row=1, column=10, value='unirrigated ratio')
        sheet2.cell(row=1, column=11, value='conventional ratio')
        sheet2.cell(row=1, column=12, value='sri ratio')
        sheet2.cell(row=1, column=13, value='high yielding ratio')
        sheet2.cell(row=1, column=14, value='local ratio')
        sheet2.cell(row=1, column=15, value='hybrid ratio')
        sheet2.cell(row=1, column=16, value='pest damage')


        while sheet2.cell(row= count2, column=4).value != None:
            #print('Entered sheet2 part')
            count2 += 1
            variety = sheet2.cell(row= count2, column=4).value
            if variety == None: break

            checker2 = False
            count = 2
            while sheet.cell(row=count, column= 2).value != None:
                count = count + 1
                #print('sheet1 entered')
                # irrigation type part below
                if variety == sheet.cell(row = count, column=8 ).value :

                    checker2 = True
                    if sheet.cell(row = count, column= 21).value.lower() == 'irrigated':
                        self.isIrrigatedType[0] += 1
                        self.isIrrigatedType_counter += 1
                    elif sheet.cell(row=count, column= 21).value.lower() == 'rainfed':
                        self.isIrrigatedType[1] += 1
                        self.isIrrigatedType_counter +=1
                    elif sheet.cell(row=count, column= 21).value.lower() == 'un-irrigated' or \
                            sheet.cell(row=count, column=21).value.lower() == 'un irrigated':
                        self.isIrrigatedType[2] +=1
                        self.isIrrigatedType_counter += 1
                    else:
                        print('Some problem in checking the irrigation type from the database')
                        print('this : ' + sheet.cell(row=count, column=21).value)

                    # seed details part1 below

                    if sheet.cell(row= count, column= 6).value.lower() == 'conventional':
                        self.seed_Details_part1[0] +=1
                        self.seed_Details_part1_counter +=1
                    elif sheet.cell(row= count, column= 6).value.lower() == 'sri':
                        self.seed_Details_part1[1] += 1
                        self.seed_Details_part1_counter += 1
                    else:
                        print('Some problem in checking the seed type part 1')

                    # seed details part2 below

                    if sheet.cell(row=count, column=7).value.lower() == 'high yielding variety':
                        self.seed_Details_part2[0] += 1
                        self.seed_Details_part2_counter += 1
                    elif sheet.cell(row=count, column=7).value.lower() == 'local':
                        self.seed_Details_part2[1] += 1
                        self.seed_Details_part2_counter += 1
                    elif sheet.cell(row=count, column=7).value.lower() == 'hybrid':
                        self.seed_Details_part2[2] += 1
                        self.seed_Details_part2_counter += 1

                    else:
                        print('Some problem in checking the seed type part 2')

                    if type(sheet.cell(row=count, column= 24).value) == float or type(sheet.cell(row=count, column=24).value) == int:
                        self.pestDamage += sheet.cell(row=count, column= 24).value
                        self.pestDamage_counter += 1

                    ## also add an extension for hybrid

                if variety != sheet.cell(row = count, column=8 ).value and checker2==True:
                    #print('reached here 11')

                    ratio = self.isIrrigatedType[0] / self.isIrrigatedType_counter
                    sheet2.cell(row= count2, column=8, value=ratio)
                    #print(ratio)
                    ratio = self.isIrrigatedType[1] / self.isIrrigatedType_counter
                    sheet2.cell(row=count2, column=9, value=ratio)
                    #print(ratio)
                    ratio = self.isIrrigatedType[2] / self.isIrrigatedType_counter
                    sheet2.cell(row=count2, column=10, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part1[0] / self.seed_Details_part1_counter
                    sheet2.cell(row = count2, column= 11, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part1[1] / self.seed_Details_part1_counter
                    sheet2.cell(row=count2, column=12, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part2[0] / self.seed_Details_part2_counter
                    sheet2.cell(row=count2, column=13, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part2[1] / self.seed_Details_part2_counter
                    sheet2.cell(row=count2, column=14, value=ratio)
                    #print(ratio)
                    ratio = self.seed_Details_part2[2] / self.seed_Details_part2_counter
                    sheet2.cell(row=count2, column=15, value=ratio)

                    try :
                        ratio = self.pestDamage / self.pestDamage_counter
                    except ZeroDivisionError :
                        sheet2.cell(row=count2, column=16, value=0)
                    else:
                        sheet2.cell(row=count2, column=16, value=ratio)

                    ratio = 0

                    # initialising all to zero

                    self.isIrrigatedType = [0,0,0]; self.isIrrigatedType_counter = 0
                    self.seed_Details_part1 = [0,0]; self.seed_Details_part1_counter = 0
                    self.seed_Details_part2 = [0,0,0]; self.seed_Details_part2_counter = 0
                    self.pestDamage_counter = 0;
                    self.pestDamage = 0

                    checker2 = False
                    break
                if sheet.cell(row=count, column=8).value == None: break
        file2.save('./ratios/ratio_good_'+self.workbook_name2+'.xlsx')




obj = Reason()
obj.individualAllotment()
obj.process_General()
obj.xx()
obj.avg_Good()
obj.bad()
