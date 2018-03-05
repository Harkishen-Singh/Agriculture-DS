from Process import Process
from openpyxl import load_workbook


class Manures(Process):
    def avg_Good(self):
        rows = 2
        file = load_workbook(self.sortedFileName_Good)
        sheet = file['Sorted']
        while (True):
            if sheet.cell(row=rows, column=2).value != None:
                rows = rows + 1
            else:
                print('came here')
                break

        total_chemfer = 0;
        c_chemfer = 0
        total_manure = 0;
        c_manure = 0
        file2 = load_workbook('./ratios/ratio_good.xlsx')
        sheet2 = file2['Sheet']
        sheet2.cell(row=1, column=5, value='ChemFer_Ratio')
        sheet2.cell(row=1, column=6, value='Manure_Ratio')
        varietyCount = 2

        for i in range(2, rows):

            if sheet.cell(row=i, column=8).value == sheet.cell(row=i + 1, column=8).value:

                if sheet.cell(row=i, column=12).value == None :
                    c_chemfer = c_chemfer + 1

                elif type(sheet.cell(row=i, column=12).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=12).value:
                        if j.isdigit():
                            number = number + j
                    total_chemfer = total_chemfer + float(number)
                    c_chemfer = c_chemfer + 1
                else:
                    total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                    c_chemfer = c_chemfer + 1

                if sheet.cell(row=i, column=11).value == None:
                    c_manure = c_manure + 1

                elif type(sheet.cell(row=i, column=11).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=11).value:
                        if j.isdigit():
                            number = number + j
                    total_manure = total_manure + float(number)
                    c_manure = c_manure + 1
                else:
                    total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                    c_manure = c_manure + 1

            '''if sheet.cell(row = i, column=8).value != sheet.cell(row = i+1, column=8).value and \
                    sheet.cell(row=i, column=8).value != sheet.cell(row=i - 1, column=8).value:
                if sheet.cell(row=i, column=12).value == None:
                    continue
                elif type(sheet.cell(row=i, column=12).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=12).value:
                        if j.isdigit():
                            number = number + j
                    total_chemfer = total_chemfer + float(number)
                    c_chemfer = c_chemfer + 1
                else:
                    total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                    c_chemfer = c_chemfer + 1
                if sheet.cell(row=i, column=11).value == None:
                    continue
                elif type(sheet.cell(row=i, column=11).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=11).value:
                        if j.isdigit():
                            number = number + j
                    total_manure = total_manure + float(number)
                    c_manure = c_manure + 1
                else:
                    total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                    c_manure = c_manure + 1
            '''

            if sheet.cell(row=i, column=8).value != sheet.cell(row=i + 1, column=8).value:

                if sheet.cell(row=i, column=12).value == None:
                    c_chemfer = c_chemfer + 1

                elif type(sheet.cell(row=i, column=12).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=12).value:
                        if j.isdigit():
                            number = number + j
                    total_chemfer = total_chemfer + float(number)
                    c_chemfer = c_chemfer + 1
                else:
                    total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                    c_chemfer = c_chemfer + 1

                if sheet.cell(row=i, column=11).value == None:
                    c_manure = c_manure + 1

                elif type(sheet.cell(row=i, column=11).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=11).value:
                        if j.isdigit():
                            number = number + j
                    total_manure = total_manure + float(number)
                    c_manure = c_manure + 1
                else:
                    total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                    c_manure = c_manure + 1

                print(str(sheet2.cell(row=varietyCount, column=4).value) + ' '+ str(total_chemfer / c_chemfer) + ' '
                      + str(total_manure / c_manure) + ' '
                      + str(sheet.cell(row= i, column=8).value))
                sheet2.cell(row=varietyCount, column=5, value=total_chemfer / c_chemfer)
                sheet2.cell(row=varietyCount, column=6, value=total_manure / c_manure)
                sheet2.cell(row=varietyCount, column=7, value=sheet.cell(row= i, column=8).value)
                varietyCount = varietyCount + 1
                total_manure = 0;
                c_manure = 0
                total_chemfer = 0;
                c_chemfer = 0

        file2.save('./ratios/ratio_good.xlsx')
        self.avg_Normal()

    def avg_Normal(self):
        rows = 2
        file = load_workbook(self.sortedFileName_Normal)
        sheet = file['Sorted']
        while (True):
            if sheet.cell(row=rows, column=2).value != None:
                rows = rows + 1
            else:
                print('came here')
                break

        total_chemfer = 0;
        c_chemfer = 0
        total_manure = 0;
        c_manure = 0
        file2 = load_workbook('./ratios/ratio_normal.xlsx')
        sheet2 = file2['Sheet']
        sheet2.cell(row=1, column=5, value='ChemFer_Ratio')
        sheet2.cell(row=1, column=6, value='Manure_Ratio')
        varietyCount = 2

        for i in range(2, rows):

            if sheet.cell(row=i, column=8).value == sheet.cell(row=i + 1, column=8).value:

                if sheet.cell(row=i, column=12).value == None :
                    c_chemfer = c_chemfer + 1

                elif type(sheet.cell(row=i, column=12).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=12).value:
                        if j.isdigit():
                            number = number + j
                    total_chemfer = total_chemfer + float(number)
                    c_chemfer = c_chemfer + 1
                else:
                    total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                    c_chemfer = c_chemfer + 1

                if sheet.cell(row=i, column=11).value == None:
                    c_manure = c_manure + 1

                elif type(sheet.cell(row=i, column=11).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=11).value:
                        if j.isdigit():
                            number = number + j
                    total_manure = total_manure + float(number)
                    c_manure = c_manure + 1
                else:
                    total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                    c_manure = c_manure + 1

            '''if sheet.cell(row = i, column=8).value != sheet.cell(row = i+1, column=8).value and \
                    sheet.cell(row=i, column=8).value != sheet.cell(row=i - 1, column=8).value:
                if sheet.cell(row=i, column=12).value == None:
                    continue
                elif type(sheet.cell(row=i, column=12).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=12).value:
                        if j.isdigit():
                            number = number + j
                    total_chemfer = total_chemfer + float(number)
                    c_chemfer = c_chemfer + 1
                else:
                    total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                    c_chemfer = c_chemfer + 1
                if sheet.cell(row=i, column=11).value == None:
                    continue
                elif type(sheet.cell(row=i, column=11).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=11).value:
                        if j.isdigit():
                            number = number + j
                    total_manure = total_manure + float(number)
                    c_manure = c_manure + 1
                else:
                    total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                    c_manure = c_manure + 1
            '''

            if sheet.cell(row=i, column=8).value != sheet.cell(row=i + 1, column=8).value:

                if sheet.cell(row=i, column=12).value == None:
                    c_chemfer = c_chemfer + 1

                elif type(sheet.cell(row=i, column=12).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=12).value:
                        if j.isdigit():
                            number = number + j
                    total_chemfer = total_chemfer + float(number)
                    c_chemfer = c_chemfer + 1
                else:
                    total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                    c_chemfer = c_chemfer + 1

                if sheet.cell(row=i, column=11).value == None:
                    c_manure = c_manure + 1

                elif type(sheet.cell(row=i, column=11).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=11).value:
                        if j.isdigit():
                            number = number + j
                    total_manure = total_manure + float(number)
                    c_manure = c_manure + 1
                else:
                    total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                    c_manure = c_manure + 1

                print(str(sheet2.cell(row=varietyCount, column=4).value) + ' '+ str(total_chemfer / c_chemfer) + ' '
                      + str(total_manure / c_manure) + ' '
                      + str(sheet.cell(row= i, column=8).value))
                sheet2.cell(row=varietyCount, column=5, value=total_chemfer / c_chemfer)
                sheet2.cell(row=varietyCount, column=6, value=total_manure / c_manure)
                sheet2.cell(row=varietyCount, column=7, value=sheet.cell(row= i, column=8).value)
                varietyCount = varietyCount + 1
                total_manure = 0;
                c_manure = 0
                total_chemfer = 0;
                c_chemfer = 0

        file2.save('./ratios/ratio_normal.xlsx')
        self.avg_Bad()

    def avg_Bad(self):
        rows = 2
        file = load_workbook(self.sortedFileName_Bad)
        sheet = file['Sorted']
        while (True):
            if sheet.cell(row=rows, column=2).value != None:
                rows = rows + 1
            else:
                print('came here')
                break

        total_chemfer = 0;
        c_chemfer = 0
        total_manure = 0;
        c_manure = 0
        file2 = load_workbook('./ratios/ratio_bad.xlsx')
        sheet2 = file2['Sheet']
        sheet2.cell(row=1, column=5, value='ChemFer_Ratio')
        sheet2.cell(row=1, column=6, value='Manure_Ratio')
        varietyCount = 2

        for i in range(2, rows):

            if sheet.cell(row=i, column=8).value == sheet.cell(row=i + 1, column=8).value:

                if sheet.cell(row=i, column=12).value == None :
                    c_chemfer = c_chemfer + 1

                elif type(sheet.cell(row=i, column=12).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=12).value:
                        if j.isdigit():
                            number = number + j
                    total_chemfer = total_chemfer + float(number)
                    c_chemfer = c_chemfer + 1
                else:
                    total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                    c_chemfer = c_chemfer + 1

                if sheet.cell(row=i, column=11).value == None:
                    c_manure = c_manure + 1

                elif type(sheet.cell(row=i, column=11).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=11).value:
                        if j.isdigit():
                            number = number + j
                    total_manure = total_manure + float(number)
                    c_manure = c_manure + 1
                else:
                    total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                    c_manure = c_manure + 1

            '''if sheet.cell(row = i, column=8).value != sheet.cell(row = i+1, column=8).value and \
                    sheet.cell(row=i, column=8).value != sheet.cell(row=i - 1, column=8).value:
                if sheet.cell(row=i, column=12).value == None:
                    continue
                elif type(sheet.cell(row=i, column=12).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=12).value:
                        if j.isdigit():
                            number = number + j
                    total_chemfer = total_chemfer + float(number)
                    c_chemfer = c_chemfer + 1
                else:
                    total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                    c_chemfer = c_chemfer + 1
                if sheet.cell(row=i, column=11).value == None:
                    continue
                elif type(sheet.cell(row=i, column=11).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=11).value:
                        if j.isdigit():
                            number = number + j
                    total_manure = total_manure + float(number)
                    c_manure = c_manure + 1
                else:
                    total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                    c_manure = c_manure + 1
            '''

            if sheet.cell(row=i, column=8).value != sheet.cell(row=i + 1, column=8).value:

                if sheet.cell(row=i, column=12).value == None:
                    c_chemfer = c_chemfer + 1

                elif type(sheet.cell(row=i, column=12).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=12).value:
                        if j.isdigit():
                            number = number + j
                    total_chemfer = total_chemfer + float(number)
                    c_chemfer = c_chemfer + 1
                else:
                    total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                    c_chemfer = c_chemfer + 1

                if sheet.cell(row=i, column=11).value == None:
                    c_manure = c_manure + 1

                elif type(sheet.cell(row=i, column=11).value) == str:
                    number = ''
                    for j in sheet.cell(row=i, column=11).value:
                        if j.isdigit():
                            number = number + j
                    total_manure = total_manure + float(number)
                    c_manure = c_manure + 1
                else:
                    total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                    c_manure = c_manure + 1

                print(str(sheet2.cell(row=varietyCount, column=4).value) + ' '+ str(total_chemfer / c_chemfer) + ' '
                      + str(total_manure / c_manure) + ' '
                      + str(sheet.cell(row= i, column=8).value))
                sheet2.cell(row=varietyCount, column=5, value=total_chemfer / c_chemfer)
                sheet2.cell(row=varietyCount, column=6, value=total_manure / c_manure)
                sheet2.cell(row=varietyCount, column=7, value=sheet.cell(row= i, column=8).value)
                varietyCount = varietyCount + 1
                total_manure = 0;
                c_manure = 0
                total_chemfer = 0;
                c_chemfer = 0

        file2.save('./ratios/ratio_bad.xlsx')

obj = Manures()
obj.individualAllotment()
obj.process_General()
obj.xx()
obj.avg_Good()