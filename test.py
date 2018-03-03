from openpyxl import Workbook

wb = Workbook() # loading the workbook
ws = wb.active # opens the workbook with the put pointer on the first sheet to write from the very first cell
ws.title = 'example of input to excel file'
# writing to the excel file

ws.cell(row = 1, column = 1, value='Harkishen Singh')
''' remember, the cell starts from 1, and not 0, unlike in the traditional array.'''

# writing multiple lines to the excel file

for i in range(1,101):
    ws.cell(row=i, column = 2, value=i**2) # writing to the excel sheet.

wb.save('sample_excel.xlsx')
wb.close()
