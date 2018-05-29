import sys, subprocess

from openpyxl import load_workbook, Workbook
from reason import Reason

class Block_Division(Reason):

    def opening(self):
        self.globalWorkBookName = self.globalWorkbookAvailable
        sortGeneralAdd = 'sorts/sortedFile_General_'+self.globalWorkBookName+'.xlsx'
        wb = load_workbook(sortGeneralAdd)
        self.sheet_general = wb['Sorted']
        self.allBlocks = []
        i = 1
        while True:
            i += 1
            if self.sheet_general.cell(row = i, column = 2).value == None:
                break
            else:
                self.allBlocks.append(self.sheet_general.cell(row = i, column = 2).value)
        sortGeneralAdd = 'sorts/sortedFile_Bad_'+self.globalWorkBookName+'.xlsx'
        wb = load_workbook(sortGeneralAdd)
        self.sheet_bad = wb['Sorted']

        sortGeneralAdd = 'sorts/sortedFile_Normal_' + self.globalWorkBookName + '.xlsx'
        wb = load_workbook(sortGeneralAdd)
        self.sheet_normal = wb['Sorted']

        sortGeneralAdd = 'sorts/sortedFile_Good_' + self.globalWorkBookName + '.xlsx'
        wb = load_workbook(sortGeneralAdd)
        self.sheet_good = wb['Sorted']



        self.blockNames = []
        self.scanner()

    def scanner(self):
        self.blockNames.append(self.allBlocks[0])

        for block in self.allBlocks :
            presence = False

            for loop in self.blockNames :
                if loop == block :
                    presence = True
                    break
            if presence == False :
                self.blockNames.append(block)
            else:
                presence = False
        print(self.blockNames)
        self.creating_required_directories()


    def creating_required_directories(self):
        defaultConsoleAdd = 'Blocks'

        # creating new directories
        print("Number of directories to be created : "+ str(len(self.blockNames)))

        for blocks in self.blockNames :

            c = subprocess.call(['mkdir', defaultConsoleAdd+'/'+blocks])
            e = subprocess.call(['mkdir', defaultConsoleAdd + '/' + blocks + '/' + 'ratios'])
            d = subprocess.call(['mkdir', defaultConsoleAdd+'/'+blocks+'/'+'sorts'])

        print('Created directories successfully ! ')

        self.extracting_required_files()

    def extracting_required_files(self):
        # creating new files


        print('Extracting files')
        i = 1
        a = 1
        b = 1
        n = 1
        g = 1

        for blocks in self.blockNames :

            nb = Workbook()
            nn = Workbook()
            ng = Workbook()
            ngeneral = Workbook()

            nsBad = nb.active
            nsBad.title = 'Sorted'

            nsNormal = nn.active
            nsNormal.title = 'Sorted'

            nsGood = ng.active
            nsGood.title = 'Sorted'

            nsGeneral = ngeneral.active
            nsGeneral.title = 'Sorted'

            i = 1
            a = 1
            b = 1
            n = 1
            g = 1

            while True:
                i += 1
                if self.sheet_general.cell(row=i, column= 2).value == None:
                    break

                if blocks == self.sheet_general.cell(row=i, column= 2).value :
                    a +=1
                    nsGeneral.cell(row=a, column=2, value=self.sheet_general.cell(row=i, column=2).value)
                    nsGeneral.cell(row=a, column=3, value=self.sheet_general.cell(row=i, column=3).value)
                    nsGeneral.cell(row=a, column=4, value=self.sheet_general.cell(row=i, column=4).value)
                    nsGeneral.cell(row=a, column=5, value=self.sheet_general.cell(row=i, column=5).value)
                    nsGeneral.cell(row=a, column=6, value=self.sheet_general.cell(row=i, column=6).value)
                    nsGeneral.cell(row=a, column=7, value=self.sheet_general.cell(row=i, column=7).value)
                    nsGeneral.cell(row=a, column=8, value=self.sheet_general.cell(row=i, column=8).value)
                    nsGeneral.cell(row=a, column=9, value=self.sheet_general.cell(row=i, column=9).value)
                    nsGeneral.cell(row=a, column=10, value=self.sheet_general.cell(row=i, column=10).value)
                    nsGeneral.cell(row=a, column=11, value=self.sheet_general.cell(row=i, column=11).value)
                    nsGeneral.cell(row=a, column=12, value=self.sheet_general.cell(row=i, column=12).value)
                    nsGeneral.cell(row=a, column=13, value=self.sheet_general.cell(row=i, column=13).value)
                    nsGeneral.cell(row=a, column=14, value=self.sheet_general.cell(row=i, column=14).value)
                    nsGeneral.cell(row=a, column=15, value=self.sheet_general.cell(row=i, column=15).value)
                    nsGeneral.cell(row=a, column=16, value=self.sheet_general.cell(row=i, column=16).value)
                    nsGeneral.cell(row=a, column=17, value=self.sheet_general.cell(row=i, column=17).value)
                    nsGeneral.cell(row=a, column=18, value=self.sheet_general.cell(row=i, column=18).value)
                    nsGeneral.cell(row=a, column=19, value=self.sheet_general.cell(row=i, column=19).value)
                    nsGeneral.cell(row=a, column=20, value=self.sheet_general.cell(row=i, column=20).value)
                    nsGeneral.cell(row=a, column=21, value=self.sheet_general.cell(row=i, column=21).value)
                    nsGeneral.cell(row=a, column=22, value=self.sheet_general.cell(row=i, column=22).value)
                    nsGeneral.cell(row=a, column=23, value=self.sheet_general.cell(row=i, column=23).value)
                    nsGeneral.cell(row=a, column=24, value=self.sheet_general.cell(row=i, column=24).value)
                    nsGeneral.cell(row=a, column=25, value=self.sheet_general.cell(row=i, column=25).value)
                    nsGeneral.cell(row=a, column=26, value=self.sheet_general.cell(row=i, column=26).value)

            i = 1
            a = 1

            while True:
                i += 1
                if self.sheet_bad.cell(row=i, column= 2).value == None:
                    break

                if blocks == self.sheet_bad.cell(row=i, column= 2).value :

                    a +=1
                    nsBad.cell(row=a, column=2, value=self.sheet_bad.cell(row=i, column=2).value)
                    nsBad.cell(row=a, column=3, value=self.sheet_bad.cell(row=i, column=3).value)
                    nsBad.cell(row=a, column=4, value=self.sheet_bad.cell(row=i, column=4).value)
                    nsBad.cell(row=a, column=5, value=self.sheet_bad.cell(row=i, column=5).value)
                    nsBad.cell(row=a, column=6, value=self.sheet_bad.cell(row=i, column=6).value)
                    nsBad.cell(row=a, column=7, value=self.sheet_bad.cell(row=i, column=7).value)
                    nsBad.cell(row=a, column=8, value=self.sheet_bad.cell(row=i, column=8).value)
                    nsBad.cell(row=a, column=9, value=self.sheet_bad.cell(row=i, column=9).value)
                    nsBad.cell(row=a, column=10, value=self.sheet_bad.cell(row=i, column=10).value)
                    nsBad.cell(row=a, column=11, value=self.sheet_bad.cell(row=i, column=11).value)
                    nsBad.cell(row=a, column=12, value=self.sheet_bad.cell(row=i, column=12).value)
                    nsBad.cell(row=a, column=13, value=self.sheet_bad.cell(row=i, column=13).value)
                    nsBad.cell(row=a, column=14, value=self.sheet_bad.cell(row=i, column=14).value)
                    nsBad.cell(row=a, column=15, value=self.sheet_bad.cell(row=i, column=15).value)
                    nsBad.cell(row=a, column=16, value=self.sheet_bad.cell(row=i, column=16).value)
                    nsBad.cell(row=a, column=17, value=self.sheet_bad.cell(row=i, column=17).value)
                    nsBad.cell(row=a, column=18, value=self.sheet_bad.cell(row=i, column=18).value)
                    nsBad.cell(row=a, column=19, value=self.sheet_bad.cell(row=i, column=19).value)
                    nsBad.cell(row=a, column=20, value=self.sheet_bad.cell(row=i, column=20).value)
                    nsBad.cell(row=a, column=21, value=self.sheet_bad.cell(row=i, column=21).value)
                    nsBad.cell(row=a, column=22, value=self.sheet_bad.cell(row=i, column=22).value)
                    nsBad.cell(row=a, column=23, value=self.sheet_bad.cell(row=i, column=23).value)
                    nsBad.cell(row=a, column=24, value=self.sheet_bad.cell(row=i, column=24).value)
                    nsBad.cell(row=a, column=25, value=self.sheet_bad.cell(row=i, column=25).value)
                    nsBad.cell(row=a, column=26, value=self.sheet_bad.cell(row=i, column=26).value)

            i = 1
            a = 1

            while True:
                i += 1
                if self.sheet_normal.cell(row=i, column= 2).value == None:
                    break

                if blocks == self.sheet_normal.cell(row=i, column= 2).value :
                    a +=1
                    nsNormal.cell(row=a, column=2, value=self.sheet_normal.cell(row=i, column=2).value)
                    nsNormal.cell(row=a, column=3, value=self.sheet_normal.cell(row=i, column=3).value)
                    nsNormal.cell(row=a, column=4, value=self.sheet_normal.cell(row=i, column=4).value)
                    nsNormal.cell(row=a, column=5, value=self.sheet_normal.cell(row=i, column=5).value)
                    nsNormal.cell(row=a, column=6, value=self.sheet_normal.cell(row=i, column=6).value)
                    nsNormal.cell(row=a, column=7, value=self.sheet_normal.cell(row=i, column=7).value)
                    nsNormal.cell(row=a, column=8, value=self.sheet_normal.cell(row=i, column=8).value)
                    nsNormal.cell(row=a, column=9, value=self.sheet_normal.cell(row=i, column=9).value)
                    nsNormal.cell(row=a, column=10, value=self.sheet_normal.cell(row=i, column=10).value)
                    nsNormal.cell(row=a, column=11, value=self.sheet_normal.cell(row=i, column=11).value)
                    nsNormal.cell(row=a, column=12, value=self.sheet_normal.cell(row=i, column=12).value)
                    nsNormal.cell(row=a, column=13, value=self.sheet_normal.cell(row=i, column=13).value)
                    nsNormal.cell(row=a, column=14, value=self.sheet_normal.cell(row=i, column=14).value)
                    nsNormal.cell(row=a, column=15, value=self.sheet_normal.cell(row=i, column=15).value)
                    nsNormal.cell(row=a, column=16, value=self.sheet_normal.cell(row=i, column=16).value)
                    nsNormal.cell(row=a, column=17, value=self.sheet_normal.cell(row=i, column=17).value)
                    nsNormal.cell(row=a, column=18, value=self.sheet_normal.cell(row=i, column=18).value)
                    nsNormal.cell(row=a, column=19, value=self.sheet_normal.cell(row=i, column=19).value)
                    nsNormal.cell(row=a, column=20, value=self.sheet_normal.cell(row=i, column=20).value)
                    nsNormal.cell(row=a, column=21, value=self.sheet_normal.cell(row=i, column=21).value)
                    nsNormal.cell(row=a, column=22, value=self.sheet_normal.cell(row=i, column=22).value)
                    nsNormal.cell(row=a, column=23, value=self.sheet_normal.cell(row=i, column=23).value)
                    nsNormal.cell(row=a, column=24, value=self.sheet_normal.cell(row=i, column=24).value)
                    nsNormal.cell(row=a, column=25, value=self.sheet_normal.cell(row=i, column=25).value)
                    nsNormal.cell(row=a, column=26, value=self.sheet_normal.cell(row=i, column=26).value)

            i = 1
            a = 1

            while True:
                i += 1
                if self.sheet_good.cell(row=i, column= 2).value == None:
                    break

                if blocks == self.sheet_good.cell(row=i, column= 2).value :
                    a +=1
                    nsGood.cell(row=a, column=2, value=self.sheet_good.cell(row=i, column=2).value)
                    nsGood.cell(row=a, column=3, value=self.sheet_good.cell(row=i, column=3).value)
                    nsGood.cell(row=a, column=4, value=self.sheet_good.cell(row=i, column=4).value)
                    nsGood.cell(row=a, column=5, value=self.sheet_good.cell(row=i, column=5).value)
                    nsGood.cell(row=a, column=6, value=self.sheet_good.cell(row=i, column=6).value)
                    nsGood.cell(row=a, column=7, value=self.sheet_good.cell(row=i, column=7).value)
                    nsGood.cell(row=a, column=8, value=self.sheet_good.cell(row=i, column=8).value)
                    nsGood.cell(row=a, column=9, value=self.sheet_good.cell(row=i, column=9).value)
                    nsGood.cell(row=a, column=10, value=self.sheet_good.cell(row=i, column=10).value)
                    nsGood.cell(row=a, column=11, value=self.sheet_good.cell(row=i, column=11).value)
                    nsGood.cell(row=a, column=12, value=self.sheet_good.cell(row=i, column=12).value)
                    nsGood.cell(row=a, column=13, value=self.sheet_good.cell(row=i, column=13).value)
                    nsGood.cell(row=a, column=14, value=self.sheet_good.cell(row=i, column=14).value)
                    nsGood.cell(row=a, column=15, value=self.sheet_good.cell(row=i, column=15).value)
                    nsGood.cell(row=a, column=16, value=self.sheet_good.cell(row=i, column=16).value)
                    nsGood.cell(row=a, column=17, value=self.sheet_good.cell(row=i, column=17).value)
                    nsGood.cell(row=a, column=18, value=self.sheet_good.cell(row=i, column=18).value)
                    nsGood.cell(row=a, column=19, value=self.sheet_good.cell(row=i, column=19).value)
                    nsGood.cell(row=a, column=20, value=self.sheet_good.cell(row=i, column=20).value)
                    nsGood.cell(row=a, column=21, value=self.sheet_good.cell(row=i, column=21).value)
                    nsGood.cell(row=a, column=22, value=self.sheet_good.cell(row=i, column=22).value)
                    nsGood.cell(row=a, column=23, value=self.sheet_good.cell(row=i, column=23).value)
                    nsGood.cell(row=a, column=24, value=self.sheet_good.cell(row=i, column=24).value)
                    nsGood.cell(row=a, column=25, value=self.sheet_good.cell(row=i, column=25).value)
                    nsGood.cell(row=a, column=26, value=self.sheet_good.cell(row=i, column=26).value)



            # saving the files

            nb.save('Blocks/'+blocks+'/sorts/'+'sortedFile_Bad_' + self.globalWorkBookName+ '_.xlsx')
            nn.save('Blocks/'+blocks+'/sorts/'+'sortedFile_Normal_' + self.globalWorkBookName+ '_.xlsx')
            ng.save('Blocks/'+blocks+'/sorts/'+'sortedFile_Good_' + self.globalWorkBookName+ '_.xlsx')
            ngeneral.save('Blocks/'+blocks+'/sorts/'+'sortedFile_General_' + self.globalWorkBookName+ '_.xlsx')
            nb.close()
            nn.close()
            ng.close()
            ngeneral.close()
            print('Required Files Created.')

        self.building_Good()

    def building_Good(self):
        print('\nBuilding "Good" Sectors ...')
        print(len(self.blockNames))
        for j in range(0, len(self.blockNames)) :


            #filename = self.sortedFileName_Good
            filename = 'Blocks/'+self.blockNames[j]+'/sorts/sortedFile_Good_'+self.globalWorkBookName+'_.xlsx'
            wb3 = load_workbook(filename)
            ws3 = wb3['Sorted']
            row = 2

            while True :
                if ws3.cell(row=row, column=2).value != None:
                    row = row + 1
                else:
                    break

            file = Workbook()
            sheet = file.active
            sheet.cell(row=1, column=1, value='greenWtProduced_ratio')
            sheet.cell(row=1, column=2, value='DryWtProduced_ratio')
            sheet.cell(row=1, column=3, value='normalYldKilo_ratio')
            sheet.cell(row=1, column=4, value='Variety_Name')

            sg = 0;
            sd = 0;
            sn = 0;
            counter = 2
            ag = 0;
            ad = 0;number = '';p1=1;p2=1
            an = 0
            checker = True;
            rowCounter = 1
            for i in range(2, row):


                    rowCounter = rowCounter + 1
                    no = i
                    if checker == True:
                        if type(ws3.cell(row=no, column=10).value) == int:

                            seedExtractor = ws3.cell(row=no, column=5).value
                            # l2 = len(seedExtractor)
                            if type(seedExtractor) == str:
                                number = ''
                                for uu in seedExtractor:
                                    if uu.isdigit() or uu == '.':
                                        number = number + uu
                                p1 = float(number)
                                number = ''

                            seedExtractor2 = ws3.cell(row=no, column=10).value
                            if type(seedExtractor2) != int:

                                for uu2 in seedExtractor2:
                                    if uu2.isdigit() or uu2 == '.': number = number + uu2
                                p2 = float(number)
                                number = ''
                                # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                            if type(seedExtractor2) == int:
                                p2 = seedExtractor2
                                # print('p1 is '+str(p1) + ' p2(int wala) is '+str(p2))

                            self.totalseeds = p1 * p2
                        else:
                            if type(ws3.cell(row=no, column=10).value) == int:

                                seedExtractor = ws3.cell(row=no, column=5).value
                                l2 = len(seedExtractor)
                                number = ''
                                for uu in seedExtractor:
                                    if uu.isdigit() or uu == '.':
                                        number = number + uu
                                p1 = float(number)
                                number = ''

                                seedExtractor2 = ws3.cell(row=no, column=10).value
                                if type(seedExtractor2) != int and type(seedExtractor2) != float:

                                    for uu2 in seedExtractor2:
                                        if uu2.isdigit() or uu2 == '.': number = number + uu2
                                    p2 = float(number)
                                    number = ''
                                    # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                                if type(seedExtractor2) == int:
                                    p2 = seedExtractor2
                                    # print('p1 is ' + str(p1) + ' p2(int wala) is ' + str(p2))

                                self.totalseeds = p1 * p2
                        # print(str(ws3.cell(row=no, column=15).value) + ' ' + str(self.totalseeds))
                        # print(str(ws3.cell(row=no, column=15).value) + ' here is it' + str(i))
                        self.greenWtProduced_ratio = ws3.cell(row=no, column=15).value / self.totalseeds
                        self.DryWtProduced_ratio = ws3.cell(row=no, column=16).value / self.totalseeds
                        self.normalYldKilo_ratio = ws3.cell(row=no, column=18).value / self.totalseeds
                        sg = sg + self.greenWtProduced_ratio
                        ag = ag + 1
                        sd = sd + self.DryWtProduced_ratio
                        ad = ad + 1
                        sn = sn + self.normalYldKilo_ratio
                        an = an + 1
                        checker = False

                    elif ws3.cell(row=no, column=8).value == ws3.cell(row=no - 1, column=8).value:
                        if type(ws3.cell(row=no, column=10).value) == int:

                            seedExtractor = ws3.cell(row=no, column=5).value
                            if type(seedExtractor) == str:
                                l2 = len(seedExtractor)
                                number = ''
                                for uu in seedExtractor:
                                    if uu.isdigit() or uu == '.':
                                        number = number + uu
                                p1 = float(number)
                                number = ''

                            seedExtractor2 = ws3.cell(row=no, column=10).value
                            if type(seedExtractor2) != int and type(seedExtractor2) != float:

                                for uu2 in seedExtractor2:
                                    if uu2.isdigit() or uu2 == '.': number = number + uu2
                                p2 = float(number)
                                number = ''
                                # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                            if type(seedExtractor2) == int:
                                p2 = seedExtractor2
                                # print('p1 is '+str(p1) + ' p2(int wala) is '+str(p2))

                            self.totalseeds = p1 * p2
                        else:
                            if type(ws3.cell(row=no, column=10).value) == int:

                                seedExtractor = ws3.cell(row=no, column=5).value
                                if type(seedExtractor) == str:
                                    l2 = len(seedExtractor)
                                    number = ''
                                    for uu in seedExtractor:
                                        if uu.isdigit() or uu == '.':
                                            number = number + uu

                                    p1 = float(number)
                                    number = ''

                                seedExtractor2 = ws3.cell(row=no, column=10).value
                                if type(seedExtractor2) != int and type(seedExtractor2) != float:

                                    for uu2 in seedExtractor2:
                                        if uu2.isdigit() or uu2 == '.': number = number + uu2
                                    p2 = float(number)
                                    number = ''
                                    # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                                if type(seedExtractor2) == int:
                                    p2 = seedExtractor2
                                    # print('p1 is ' + str(p1) + ' p2(int wala) is ' + str(p2))

                                self.totalseeds = p1 * p2
                        # print(str(ws3.cell(row=no, column=15).value) + ' ' + str(self.totalseeds))

                        self.greenWtProduced_ratio = ws3.cell(row=no, column=15).value / self.totalseeds
                        self.DryWtProduced_ratio = ws3.cell(row=no, column=16).value / self.totalseeds
                        self.normalYldKilo_ratio = ws3.cell(row=no, column=18).value / self.totalseeds
                        sg = sg + self.greenWtProduced_ratio;
                        ag = ag + 1
                        sd = sd + self.DryWtProduced_ratio;
                        ad = ad + 1
                        sn = sn + self.normalYldKilo_ratio;
                        an = an + 1

                    if ws3.cell(row=no, column=8).value != ws3.cell(row=no - 1, column=8).value:
                        sheet.cell(row=counter, column=1, value=sg / ag);
                        ag = 0;
                        sg = 0
                        sheet.cell(row=counter, column=2, value=sd / ad);
                        ad = 0;
                        sd = 0
                        sheet.cell(row=counter, column=3, value=sn / an);
                        an = 0;
                        sn = 0
                        sheet.cell(row=counter, column=4, value=ws3.cell(row=no, column=8).value)
                        checker = True
                        counter = counter + 1
            print(i)
            file.save('Blocks/'+self.blockNames[j]+'/ratios/ratio_good_' + self.globalWorkBookName + '.xlsx')
            file.close()

        self.building_Normal()

    def building_Normal(self):

        print('\nBuilding "Normal" Sectors ...')

        for j in range(0, len(self.blockNames)) :

            #filename = self.sortedFileName_Good
            filename = 'Blocks/'+self.blockNames[j]+'/sorts/sortedFile_Normal_'+self.globalWorkBookName+'_.xlsx'
            wb3 = load_workbook(filename)
            ws3 = wb3['Sorted']
            row = 2

            while (row):
                if ws3.cell(row=row, column=2).value != None:
                    row = row + 1
                else:
                    break

            file = Workbook()
            sheet = file.active
            sheet.cell(row=1, column=1, value='greenWtProduced_ratio')
            sheet.cell(row=1, column=2, value='DryWtProduced_ratio')
            sheet.cell(row=1, column=3, value='normalYldKilo_ratio')
            sheet.cell(row=1, column=4, value='Variety_Name')

            sg = 0;
            sd = 0;
            sn = 0;
            counter = 2
            ag = 0;
            ad = 0;number = '';p1=1;p2=1
            an = 0
            checker = True;
            rowCounter = 1
            for i in range(2, row):


                    rowCounter = rowCounter + 1
                    no = i
                    if checker == True:
                        if type(ws3.cell(row=no, column=10).value) == int:

                            seedExtractor = ws3.cell(row=no, column=5).value
                            # l2 = len(seedExtractor)
                            if type(seedExtractor) == str:
                                number = ''
                                for uu in seedExtractor:
                                    if uu.isdigit() or uu == '.':
                                        number = number + uu
                                p1 = float(number)
                                number = ''

                            seedExtractor2 = ws3.cell(row=no, column=10).value
                            if type(seedExtractor2) != int:

                                for uu2 in seedExtractor2:
                                    if uu2.isdigit() or uu2 == '.': number = number + uu2
                                p2 = float(number)
                                number = ''
                                # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                            if type(seedExtractor2) == int:
                                p2 = seedExtractor2
                                # print('p1 is '+str(p1) + ' p2(int wala) is '+str(p2))

                            self.totalseeds = p1 * p2
                        else:
                            if type(ws3.cell(row=no, column=10).value) == int:

                                seedExtractor = ws3.cell(row=no, column=5).value
                                l2 = len(seedExtractor)
                                number = ''
                                for uu in seedExtractor:
                                    if uu.isdigit() or uu == '.':
                                        number = number + uu
                                p1 = float(number)
                                number = ''

                                seedExtractor2 = ws3.cell(row=no, column=10).value
                                if type(seedExtractor2) != int and type(seedExtractor2) != float:

                                    for uu2 in seedExtractor2:
                                        if uu2.isdigit() or uu2 == '.': number = number + uu2
                                    p2 = float(number)
                                    number = ''
                                    # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                                if type(seedExtractor2) == int:
                                    p2 = seedExtractor2
                                    # print('p1 is ' + str(p1) + ' p2(int wala) is ' + str(p2))

                                self.totalseeds = p1 * p2
                        # print(str(ws3.cell(row=no, column=15).value) + ' ' + str(self.totalseeds))
                        # print(str(ws3.cell(row=no, column=15).value) + ' here is it' + str(i))
                        self.greenWtProduced_ratio = ws3.cell(row=no, column=15).value / self.totalseeds
                        self.DryWtProduced_ratio = ws3.cell(row=no, column=16).value / self.totalseeds
                        self.normalYldKilo_ratio = ws3.cell(row=no, column=18).value / self.totalseeds
                        sg = sg + self.greenWtProduced_ratio
                        ag = ag + 1
                        sd = sd + self.DryWtProduced_ratio
                        ad = ad + 1
                        sn = sn + self.normalYldKilo_ratio
                        an = an + 1
                        checker = False

                    elif ws3.cell(row=no, column=8).value == ws3.cell(row=no - 1, column=8).value:
                        if type(ws3.cell(row=no, column=10).value) == int:

                            seedExtractor = ws3.cell(row=no, column=5).value
                            if type(seedExtractor) == str:
                                l2 = len(seedExtractor)
                                number = ''
                                for uu in seedExtractor:
                                    if uu.isdigit() or uu == '.':
                                        number = number + uu
                                p1 = float(number)
                                number = ''

                            seedExtractor2 = ws3.cell(row=no, column=10).value
                            if type(seedExtractor2) != int and type(seedExtractor2) != float:

                                for uu2 in seedExtractor2:
                                    if uu2.isdigit() or uu2 == '.': number = number + uu2
                                p2 = float(number)
                                number = ''
                                # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                            if type(seedExtractor2) == int:
                                p2 = seedExtractor2
                                # print('p1 is '+str(p1) + ' p2(int wala) is '+str(p2))

                            self.totalseeds = p1 * p2
                        else:
                            if type(ws3.cell(row=no, column=10).value) == int:

                                seedExtractor = ws3.cell(row=no, column=5).value
                                if type(seedExtractor) == str:
                                    l2 = len(seedExtractor)
                                    number = ''
                                    for uu in seedExtractor:
                                        if uu.isdigit() or uu == '.':
                                            number = number + uu

                                    p1 = float(number)
                                    number = ''

                                seedExtractor2 = ws3.cell(row=no, column=10).value
                                if type(seedExtractor2) != int and type(seedExtractor2) != float:

                                    for uu2 in seedExtractor2:
                                        if uu2.isdigit() or uu2 == '.': number = number + uu2
                                    p2 = float(number)
                                    number = ''
                                    # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                                if type(seedExtractor2) == int:
                                    p2 = seedExtractor2
                                    # print('p1 is ' + str(p1) + ' p2(int wala) is ' + str(p2))

                                self.totalseeds = p1 * p2
                        # print(str(ws3.cell(row=no, column=15).value) + ' ' + str(self.totalseeds))

                        self.greenWtProduced_ratio = ws3.cell(row=no, column=15).value / self.totalseeds
                        self.DryWtProduced_ratio = ws3.cell(row=no, column=16).value / self.totalseeds
                        self.normalYldKilo_ratio = ws3.cell(row=no, column=18).value / self.totalseeds
                        sg = sg + self.greenWtProduced_ratio;
                        ag = ag + 1
                        sd = sd + self.DryWtProduced_ratio;
                        ad = ad + 1
                        sn = sn + self.normalYldKilo_ratio;
                        an = an + 1

                    if ws3.cell(row=no, column=8).value != ws3.cell(row=no - 1, column=8).value:
                        sheet.cell(row=counter, column=1, value=sg / ag);
                        ag = 0;
                        sg = 0
                        sheet.cell(row=counter, column=2, value=sd / ad);
                        ad = 0;
                        sd = 0
                        sheet.cell(row=counter, column=3, value=sn / an);
                        an = 0;
                        sn = 0
                        sheet.cell(row=counter, column=4, value=ws3.cell(row=no, column=8).value)
                        checker = True
                        counter = counter + 1

            file.save('Blocks/'+self.blockNames[j]+'/ratios/ratio_normal_' + self.globalWorkBookName + '.xlsx')
            file.close()
        self.building_Bad()

    def building_Bad(self):

        print('\nBuilding "Bad" Sectors ...\n')
        for j in range(0, len(self.blockNames)) :


            #filename = self.sortedFileName_Good
            filename = 'Blocks/'+self.blockNames[j]+'/sorts/sortedFile_Bad_'+self.globalWorkBookName+'_.xlsx'
            wb3 = load_workbook(filename)
            ws3 = wb3['Sorted']
            row = 2

            while (row):
                if ws3.cell(row=row, column=2).value != None:
                    row = row + 1
                else:
                    break

            file = Workbook()
            sheet = file.active
            sheet.cell(row=1, column=1, value='greenWtProduced_ratio')
            sheet.cell(row=1, column=2, value='DryWtProduced_ratio')
            sheet.cell(row=1, column=3, value='normalYldKilo_ratio')
            sheet.cell(row=1, column=4, value='Variety_Name')

            sg = 0;
            sd = 0;
            sn = 0;
            counter = 2
            ag = 0;
            ad = 0;number = '';p1=1;p2=1
            an = 0
            checker = True;
            rowCounter = 1
            for i in range(2, row):


                    rowCounter = rowCounter + 1
                    no = i
                    if checker == True:
                        if type(ws3.cell(row=no, column=10).value) == int:

                            seedExtractor = ws3.cell(row=no, column=5).value
                            # l2 = len(seedExtractor)
                            if type(seedExtractor) == str:
                                number = ''
                                for uu in seedExtractor:
                                    if uu.isdigit() or uu == '.':
                                        number = number + uu
                                p1 = float(number)
                                number = ''

                            seedExtractor2 = ws3.cell(row=no, column=10).value
                            if type(seedExtractor2) != int:

                                for uu2 in seedExtractor2:
                                    if uu2.isdigit() or uu2 == '.': number = number + uu2
                                p2 = float(number)
                                number = ''
                                # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                            if type(seedExtractor2) == int:
                                p2 = seedExtractor2
                                # print('p1 is '+str(p1) + ' p2(int wala) is '+str(p2))

                            self.totalseeds = p1 * p2
                        else:
                            if type(ws3.cell(row=no, column=10).value) == int:

                                seedExtractor = ws3.cell(row=no, column=5).value
                                l2 = len(seedExtractor)
                                number = ''
                                for uu in seedExtractor:
                                    if uu.isdigit() or uu == '.':
                                        number = number + uu
                                p1 = float(number)
                                number = ''

                                seedExtractor2 = ws3.cell(row=no, column=10).value
                                if type(seedExtractor2) != int and type(seedExtractor2) != float:

                                    for uu2 in seedExtractor2:
                                        if uu2.isdigit() or uu2 == '.': number = number + uu2
                                    p2 = float(number)
                                    number = ''
                                    # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                                if type(seedExtractor2) == int:
                                    p2 = seedExtractor2
                                    # print('p1 is ' + str(p1) + ' p2(int wala) is ' + str(p2))

                                self.totalseeds = p1 * p2
                        # print(str(ws3.cell(row=no, column=15).value) + ' ' + str(self.totalseeds))
                        # print(str(ws3.cell(row=no, column=15).value) + ' here is it' + str(i))
                        self.greenWtProduced_ratio = ws3.cell(row=no, column=15).value / self.totalseeds
                        self.DryWtProduced_ratio = ws3.cell(row=no, column=16).value / self.totalseeds
                        self.normalYldKilo_ratio = ws3.cell(row=no, column=18).value / self.totalseeds
                        sg = sg + self.greenWtProduced_ratio
                        ag = ag + 1
                        sd = sd + self.DryWtProduced_ratio
                        ad = ad + 1
                        sn = sn + self.normalYldKilo_ratio
                        an = an + 1
                        checker = False

                    elif ws3.cell(row=no, column=8).value == ws3.cell(row=no - 1, column=8).value:
                        if type(ws3.cell(row=no, column=10).value) == int:

                            seedExtractor = ws3.cell(row=no, column=5).value
                            if type(seedExtractor) == str:
                                l2 = len(seedExtractor)
                                number = ''
                                for uu in seedExtractor:
                                    if uu.isdigit() or uu == '.':
                                        number = number + uu
                                p1 = float(number)
                                number = ''

                            seedExtractor2 = ws3.cell(row=no, column=10).value
                            if type(seedExtractor2) != int and type(seedExtractor2) != float:

                                for uu2 in seedExtractor2:
                                    if uu2.isdigit() or uu2 == '.': number = number + uu2
                                p2 = float(number)
                                number = ''
                                # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                            if type(seedExtractor2) == int:
                                p2 = seedExtractor2
                                # print('p1 is '+str(p1) + ' p2(int wala) is '+str(p2))

                            self.totalseeds = p1 * p2
                        else:
                            if type(ws3.cell(row=no, column=10).value) == int:

                                seedExtractor = ws3.cell(row=no, column=5).value
                                if type(seedExtractor) == str:
                                    l2 = len(seedExtractor)
                                    number = ''
                                    for uu in seedExtractor:
                                        if uu.isdigit() or uu == '.':
                                            number = number + uu

                                    p1 = float(number)
                                    number = ''

                                seedExtractor2 = ws3.cell(row=no, column=10).value
                                if type(seedExtractor2) != int and type(seedExtractor2) != float:

                                    for uu2 in seedExtractor2:
                                        if uu2.isdigit() or uu2 == '.': number = number + uu2
                                    p2 = float(number)
                                    number = ''
                                    # print('p1 is ' + str(p1) + ' p2 is ' + str(p2))
                                if type(seedExtractor2) == int:
                                    p2 = seedExtractor2
                                    # print('p1 is ' + str(p1) + ' p2(int wala) is ' + str(p2))

                                self.totalseeds = p1 * p2
                        # print(str(ws3.cell(row=no, column=15).value) + ' ' + str(self.totalseeds))

                        self.greenWtProduced_ratio = ws3.cell(row=no, column=15).value / self.totalseeds
                        self.DryWtProduced_ratio = ws3.cell(row=no, column=16).value / self.totalseeds
                        self.normalYldKilo_ratio = ws3.cell(row=no, column=18).value / self.totalseeds
                        sg = sg + self.greenWtProduced_ratio;
                        ag = ag + 1
                        sd = sd + self.DryWtProduced_ratio;
                        ad = ad + 1
                        sn = sn + self.normalYldKilo_ratio;
                        an = an + 1

                    if ws3.cell(row=no, column=8).value != ws3.cell(row=no - 1, column=8).value:
                        sheet.cell(row=counter, column=1, value=sg / ag);
                        ag = 0;
                        sg = 0
                        sheet.cell(row=counter, column=2, value=sd / ad);
                        ad = 0;
                        sd = 0
                        sheet.cell(row=counter, column=3, value=sn / an);
                        an = 0;
                        sn = 0
                        sheet.cell(row=counter, column=4, value=ws3.cell(row=no, column=8).value)
                        checker = True
                        counter = counter + 1

            file.save('Blocks/'+self.blockNames[j]+'/ratios/ratio_bad_' + self.globalWorkBookName + '.xlsx')
            file.close()
        self.manures_good()


    def manures_good(self):
        print('Calculating individual yeilds. This might take some time.')
        for jj in range(0,len(self.blockNames)):
            rows = 2
            file = load_workbook('Blocks/'+self.blockNames[jj]+'/sorts/sortedFile_Good_' + self.globalWorkBookName + '_.xlsx')
            sheet = file['Sorted']
            while (True):
                if sheet.cell(row=rows, column=2).value != None:
                    rows = rows + 1
                else:

                    break

            total_chemfer = 0;
            c_chemfer = 0
            total_manure = 0;
            c_manure = 0
            file2 = load_workbook('Blocks/'+self.blockNames[jj]+'/ratios/ratio_good_' + self.globalWorkBookName + '.xlsx')
            sheet2 = file2['Sheet']
            sheet2.cell(row=1, column=5, value='ChemFer_Ratio')
            sheet2.cell(row=1, column=6, value='Manure_Ratio')
            varietyCount = 2

            for i in range(2, rows):

                if sheet.cell(row=i, column=8).value == sheet.cell(row=i + 1, column=8).value:

                    if sheet.cell(row=i, column=12).value == None:
                        c_chemfer = c_chemfer + 1

                    elif type(sheet.cell(row=i, column=12).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=12).value:
                            if j.isdigit():
                                number = number + j
                        total_chemfer = total_chemfer + float(number)
                        c_chemfer = c_chemfer + 1
                    else:
                        total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                        c_chemfer = c_chemfer + 1

                    if sheet.cell(row=i, column=11).value == None:
                        c_manure = c_manure + 1

                    elif type(sheet.cell(row=i, column=11).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=11).value:
                            if j.isdigit():
                                number = number + j
                        total_manure = total_manure + float(number)
                        c_manure = c_manure + 1
                    else:
                        total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                        c_manure = c_manure + 1

                if sheet.cell(row=i, column=8).value != sheet.cell(row=i + 1, column=8).value:

                    if sheet.cell(row=i, column=12).value == None:
                        c_chemfer = c_chemfer + 1

                    elif type(sheet.cell(row=i, column=12).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=12).value:
                            if j.isdigit():
                                number = number + j
                        total_chemfer = total_chemfer + float(number)
                        c_chemfer = c_chemfer + 1
                    else:
                        total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                        c_chemfer = c_chemfer + 1

                    if sheet.cell(row=i, column=11).value == None:
                        c_manure = c_manure + 1

                    elif type(sheet.cell(row=i, column=11).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=11).value:
                            if j.isdigit():
                                number = number + j
                        total_manure = total_manure + float(number)
                        c_manure = c_manure + 1
                    else:
                        total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                        c_manure = c_manure + 1

                    '''print(str(sheet2.cell(row=varietyCount, column=4).value) + ' '+ str(total_chemfer / c_chemfer) + ' '
                          + str(total_manure / c_manure) + ' '
                          + str(sheet.cell(row= i, column=8).value))
                    '''
                    sheet2.cell(row=varietyCount, column=5, value=total_chemfer / c_chemfer)
                    sheet2.cell(row=varietyCount, column=6, value=total_manure / c_manure)
                    sheet2.cell(row=varietyCount, column=7, value=sheet.cell(row=i, column=8).value)
                    varietyCount = varietyCount + 1
                    total_manure = 0;
                    c_manure = 0
                    total_chemfer = 0;
                    c_chemfer = 0

            file2.save('Blocks/'+self.blockNames[jj]+'/ratios/ratio_good_' + self.globalWorkBookName + '.xlsx')
            file2.close()
            self.manures_normal()

    def manures_normal(self):
        for jj in range(0,len(self.blockNames)):
            rows = 2
            file = load_workbook('Blocks/'+self.blockNames[jj]+'/sorts/sortedFile_Normal_' + self.globalWorkBookName + '_.xlsx')
            sheet = file['Sorted']
            while (True):
                if sheet.cell(row=rows, column=2).value != None:
                    rows = rows + 1
                else:

                    break

            total_chemfer = 0;
            c_chemfer = 0
            total_manure = 0;
            c_manure = 0
            file2 = load_workbook('Blocks/'+self.blockNames[jj]+'/ratios/ratio_normal_' + self.globalWorkBookName + '.xlsx')
            sheet2 = file2['Sheet']
            sheet2.cell(row=1, column=5, value='ChemFer_Ratio')
            sheet2.cell(row=1, column=6, value='Manure_Ratio')
            varietyCount = 2

            for i in range(2, rows):

                if sheet.cell(row=i, column=8).value == sheet.cell(row=i + 1, column=8).value:

                    if sheet.cell(row=i, column=12).value == None:
                        c_chemfer = c_chemfer + 1

                    elif type(sheet.cell(row=i, column=12).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=12).value:
                            if j.isdigit():
                                number = number + j
                        total_chemfer = total_chemfer + float(number)
                        c_chemfer = c_chemfer + 1
                    else:
                        total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                        c_chemfer = c_chemfer + 1

                    if sheet.cell(row=i, column=11).value == None:
                        c_manure = c_manure + 1

                    elif type(sheet.cell(row=i, column=11).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=11).value:
                            if j.isdigit():
                                number = number + j
                        total_manure = total_manure + float(number)
                        c_manure = c_manure + 1
                    else:
                        total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                        c_manure = c_manure + 1

                if sheet.cell(row=i, column=8).value != sheet.cell(row=i + 1, column=8).value:

                    if sheet.cell(row=i, column=12).value == None:
                        c_chemfer = c_chemfer + 1

                    elif type(sheet.cell(row=i, column=12).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=12).value:
                            if j.isdigit():
                                number = number + j
                        total_chemfer = total_chemfer + float(number)
                        c_chemfer = c_chemfer + 1
                    else:
                        total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                        c_chemfer = c_chemfer + 1

                    if sheet.cell(row=i, column=11).value == None:
                        c_manure = c_manure + 1

                    elif type(sheet.cell(row=i, column=11).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=11).value:
                            if j.isdigit():
                                number = number + j
                        total_manure = total_manure + float(number)
                        c_manure = c_manure + 1
                    else:
                        total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                        c_manure = c_manure + 1

                    '''print(str(sheet2.cell(row=varietyCount, column=4).value) + ' '+ str(total_chemfer / c_chemfer) + ' '
                          + str(total_manure / c_manure) + ' '
                          + str(sheet.cell(row= i, column=8).value))
                    '''
                    sheet2.cell(row=varietyCount, column=5, value=total_chemfer / c_chemfer)
                    sheet2.cell(row=varietyCount, column=6, value=total_manure / c_manure)
                    sheet2.cell(row=varietyCount, column=7, value=sheet.cell(row=i, column=8).value)
                    varietyCount = varietyCount + 1
                    total_manure = 0;
                    c_manure = 0
                    total_chemfer = 0;
                    c_chemfer = 0

            file2.save('Blocks/'+self.blockNames[jj]+'/ratios/ratio_normal_' + self.globalWorkBookName + '.xlsx')
            file2.close()
        self.manures_bad()
    def manures_bad(self):
        for jj in range(0,len(self.blockNames)):
            rows = 2
            file = load_workbook('Blocks/'+self.blockNames[jj]+'/sorts/sortedFile_Bad_' + self.globalWorkBookName + '_.xlsx')
            sheet = file['Sorted']
            while (True):
                if sheet.cell(row=rows, column=2).value != None:
                    rows = rows + 1
                else:

                    break

            total_chemfer = 0;
            c_chemfer = 0
            total_manure = 0;
            c_manure = 0
            file2 = load_workbook('Blocks/'+self.blockNames[jj]+'/ratios/ratio_bad_' + self.globalWorkBookName + '.xlsx')
            sheet2 = file2['Sheet']
            sheet2.cell(row=1, column=5, value='ChemFer_Ratio')
            sheet2.cell(row=1, column=6, value='Manure_Ratio')
            varietyCount = 2

            for i in range(2, rows):

                if sheet.cell(row=i, column=8).value == sheet.cell(row=i + 1, column=8).value:

                    if sheet.cell(row=i, column=12).value == None:
                        c_chemfer = c_chemfer + 1

                    elif type(sheet.cell(row=i, column=12).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=12).value:
                            if j.isdigit():
                                number = number + j
                        total_chemfer = total_chemfer + float(number)
                        c_chemfer = c_chemfer + 1
                    else:
                        total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                        c_chemfer = c_chemfer + 1

                    if sheet.cell(row=i, column=11).value == None:
                        c_manure = c_manure + 1

                    elif type(sheet.cell(row=i, column=11).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=11).value:
                            if j.isdigit():
                                number = number + j
                        total_manure = total_manure + float(number)
                        c_manure = c_manure + 1
                    else:
                        total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                        c_manure = c_manure + 1

                if sheet.cell(row=i, column=8).value != sheet.cell(row=i + 1, column=8).value:

                    if sheet.cell(row=i, column=12).value == None:
                        c_chemfer = c_chemfer + 1

                    elif type(sheet.cell(row=i, column=12).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=12).value:
                            if j.isdigit():
                                number = number + j
                        total_chemfer = total_chemfer + float(number)
                        c_chemfer = c_chemfer + 1
                    else:
                        total_chemfer = total_chemfer + float(sheet.cell(row=i, column=12).value)
                        c_chemfer = c_chemfer + 1

                    if sheet.cell(row=i, column=11).value == None:
                        c_manure = c_manure + 1

                    elif type(sheet.cell(row=i, column=11).value) == str:
                        number = ''
                        for j in sheet.cell(row=i, column=11).value:
                            if j.isdigit():
                                number = number + j
                        total_manure = total_manure + float(number)
                        c_manure = c_manure + 1
                    else:
                        total_manure = total_manure + float(sheet.cell(row=i, column=11).value)
                        c_manure = c_manure + 1

                    '''print(str(sheet2.cell(row=varietyCount, column=4).value) + ' '+ str(total_chemfer / c_chemfer) + ' '
                          + str(total_manure / c_manure) + ' '
                          + str(sheet.cell(row= i, column=8).value))
                    '''
                    sheet2.cell(row=varietyCount, column=5, value=total_chemfer / c_chemfer)
                    sheet2.cell(row=varietyCount, column=6, value=total_manure / c_manure)
                    sheet2.cell(row=varietyCount, column=7, value=sheet.cell(row=i, column=8).value)
                    varietyCount = varietyCount + 1
                    total_manure = 0;
                    c_manure = 0
                    total_chemfer = 0;
                    c_chemfer = 0

            file2.save('Blocks/'+self.blockNames[jj]+'/ratios/ratio_bad_' + self.globalWorkBookName + '.xlsx')
            file2.close()
            file.close()
        self.reason_bad()

    def reason_bad(self):
        for jj in range(0, len(self.blockNames)):
            file = load_workbook('Blocks/'+self.blockNames[jj]+'/sorts/sortedFile_Bad_' + self.globalWorkBookName + '_.xlsx')
            sheet = file['Sorted']
            self.isIrrigatedType = [0] * 3
            self.isIrrigatedType_counter = 0

            # irrigated, rainfed, unirrigated
            self.seed_Details_part1 = [0] * 2
            self.seed_Details_part1_counter = 0

            # conventional , sri
            self.seed_Details_part2 = [0] * 3
            self.seed_Details_part2_counter = 0

            # high yielding, local, hybrid
            count = 1

            # pests damage avergae calculation
            self.pestDamage = 0
            self.pestDamage_counter = 0

            file2 = load_workbook('Blocks/'+self.blockNames[jj]+'/ratios/ratio_bad_' + self.globalWorkBookName + '.xlsx')
            sheet2 = file2['Sheet']
            count2 = 1
            sheet2.cell(row=1, column=8, value='irrigated ratio')
            sheet2.cell(row=1, column=9, value='rainfed ratio')
            sheet2.cell(row=1, column=10, value='unirrigated ratio')
            sheet2.cell(row=1, column=11, value='conventional ratio')
            sheet2.cell(row=1, column=12, value='sri ratio')
            sheet2.cell(row=1, column=13, value='high yielding ratio')
            sheet2.cell(row=1, column=14, value='local ratio')
            sheet2.cell(row=1, column=15, value='hybrid ratio')
            sheet2.cell(row=1, column=16, value='pest damage')

            while sheet2.cell(row=count2, column=4).value != None:
                # print('Entered sheet2 part')
                count2 += 1
                variety = sheet2.cell(row=count2, column=4).value
                if variety == None: break

                checker2 = False
                count = 2
                while sheet.cell(row=count, column=2).value != None:
                    count = count + 1
                    # print('sheet1 entered')
                    # irrigation type part below
                    if variety == sheet.cell(row=count, column=8).value:

                        checker2 = True
                        if sheet.cell(row=count, column=21).value.lower() == 'irrigated':
                            self.isIrrigatedType[0] += 1
                            self.isIrrigatedType_counter += 1
                        elif sheet.cell(row=count, column=21).value.lower() == 'rainfed':
                            self.isIrrigatedType[1] += 1
                            self.isIrrigatedType_counter += 1
                        elif sheet.cell(row=count, column=21).value.lower() == 'un-irrigated' or \
                                sheet.cell(row=count, column=21).value.lower() == 'un irrigated':
                            self.isIrrigatedType[2] += 1
                            self.isIrrigatedType_counter += 1
                        else:
                            print('Some problem in checking the irrigation type from the database')
                            print('this : ' + sheet.cell(row=count, column=21).value)

                        # seed details part1 below

                        if sheet.cell(row=count, column=6).value.lower() == 'conventional':
                            self.seed_Details_part1[0] += 1
                            self.seed_Details_part1_counter += 1
                        elif sheet.cell(row=count, column=6).value.lower() == 'sri':
                            self.seed_Details_part1[1] += 1
                            self.seed_Details_part1_counter += 1
                        else:
                            print('Some problem in checking the seed type part 1')

                        # seed details part2 below

                        if sheet.cell(row=count, column=7).value.lower() == 'high yielding variety':
                            self.seed_Details_part2[0] += 1
                            self.seed_Details_part2_counter += 1
                        elif sheet.cell(row=count, column=7).value.lower() == 'local':
                            self.seed_Details_part2[1] += 1
                            self.seed_Details_part2_counter += 1
                        elif sheet.cell(row=count, column=7).value.lower() == 'hybrid':
                            self.seed_Details_part2[2] += 1
                            self.seed_Details_part2_counter += 1

                        else:
                            print('Some problem in checking the seed type part 2')

                        if type(sheet.cell(row=count, column=24).value) == float or type(
                                sheet.cell(row=count, column=24).value) == int:
                            self.pestDamage += sheet.cell(row=count, column=24).value
                            self.pestDamage_counter += 1

                    if variety != sheet.cell(row=count, column=8).value and checker2 == True:
                        # print('reached here 11')

                        ratio = self.isIrrigatedType[0] / self.isIrrigatedType_counter
                        sheet2.cell(row=count2, column=8, value=ratio)
                        # print(ratio)
                        ratio = self.isIrrigatedType[1] / self.isIrrigatedType_counter
                        sheet2.cell(row=count2, column=9, value=ratio)
                        # print(ratio)
                        ratio = self.isIrrigatedType[2] / self.isIrrigatedType_counter
                        sheet2.cell(row=count2, column=10, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part1[0] / self.seed_Details_part1_counter
                        sheet2.cell(row=count2, column=11, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part1[1] / self.seed_Details_part1_counter
                        sheet2.cell(row=count2, column=12, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part2[0] / self.seed_Details_part2_counter
                        sheet2.cell(row=count2, column=13, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part2[1] / self.seed_Details_part2_counter
                        sheet2.cell(row=count2, column=14, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part2[2] / self.seed_Details_part2_counter
                        sheet2.cell(row=count2, column=15, value=ratio)
                        try:
                            ratio = self.pestDamage / self.pestDamage_counter
                        except ZeroDivisionError:
                            sheet2.cell(row=count2, column=16, value=0)
                        else:
                            sheet2.cell(row=count2, column=16, value=ratio)

                        ratio = 0
                        # initialising all to zero

                        self.isIrrigatedType = [0, 0, 0];
                        self.isIrrigatedType_counter = 0
                        self.seed_Details_part1 = [0, 0];
                        self.seed_Details_part1_counter = 0
                        self.seed_Details_part2 = [0, 0, 0];
                        self.seed_Details_part2_counter = 0
                        self.pestDamage_counter = 0;
                        self.pestDamage = 0
                        checker2 = False
                        break
                    if sheet.cell(row=count, column=8).value == None: break
            file2.save('Blocks/'+self.blockNames[jj]+'/ratios/ratio_bad_' + self.globalWorkBookName + '.xlsx')
            file2.close()
            file.close()
        self.reason_normal()

    def reason_normal(self):
        for jj in range(0, len(self.blockNames)):
            file = load_workbook('Blocks/'+self.blockNames[jj]+'/sorts/sortedFile_Normal_' + self.globalWorkBookName + '_.xlsx')
            sheet = file['Sorted']
            self.isIrrigatedType = [0] * 3
            self.isIrrigatedType_counter = 0

            # irrigated, rainfed, unirrigated
            self.seed_Details_part1 = [0] * 2
            self.seed_Details_part1_counter = 0

            # conventional , sri
            self.seed_Details_part2 = [0] * 3
            self.seed_Details_part2_counter = 0

            # high yielding, local, hybrid
            count = 1

            # pests damage avergae calculation
            self.pestDamage = 0
            self.pestDamage_counter = 0

            file2 = load_workbook('Blocks/'+self.blockNames[jj]+'/ratios/ratio_normal_' + self.globalWorkBookName + '.xlsx')
            sheet2 = file2['Sheet']
            count2 = 1
            sheet2.cell(row=1, column=8, value='irrigated ratio')
            sheet2.cell(row=1, column=9, value='rainfed ratio')
            sheet2.cell(row=1, column=10, value='unirrigated ratio')
            sheet2.cell(row=1, column=11, value='conventional ratio')
            sheet2.cell(row=1, column=12, value='sri ratio')
            sheet2.cell(row=1, column=13, value='high yielding ratio')
            sheet2.cell(row=1, column=14, value='local ratio')
            sheet2.cell(row=1, column=15, value='hybrid ratio')
            sheet2.cell(row=1, column=16, value='pest damage')

            while sheet2.cell(row=count2, column=4).value != None:
                # print('Entered sheet2 part')
                count2 += 1
                variety = sheet2.cell(row=count2, column=4).value
                if variety == None: break

                checker2 = False
                count = 2
                while sheet.cell(row=count, column=2).value != None:
                    count = count + 1
                    # print('sheet1 entered')
                    # irrigation type part below
                    if variety == sheet.cell(row=count, column=8).value:

                        checker2 = True
                        if sheet.cell(row=count, column=21).value.lower() == 'irrigated':
                            self.isIrrigatedType[0] += 1
                            self.isIrrigatedType_counter += 1
                        elif sheet.cell(row=count, column=21).value.lower() == 'rainfed':
                            self.isIrrigatedType[1] += 1
                            self.isIrrigatedType_counter += 1
                        elif sheet.cell(row=count, column=21).value.lower() == 'un-irrigated' or \
                                sheet.cell(row=count, column=21).value.lower() == 'un irrigated':
                            self.isIrrigatedType[2] += 1
                            self.isIrrigatedType_counter += 1
                        else:
                            print('Some problem in checking the irrigation type from the database')
                            print('this : ' + sheet.cell(row=count, column=21).value)

                        # seed details part1 below

                        if sheet.cell(row=count, column=6).value.lower() == 'conventional':
                            self.seed_Details_part1[0] += 1
                            self.seed_Details_part1_counter += 1
                        elif sheet.cell(row=count, column=6).value.lower() == 'sri':
                            self.seed_Details_part1[1] += 1
                            self.seed_Details_part1_counter += 1
                        else:
                            print('Some problem in checking the seed type part 1')

                        # seed details part2 below

                        if sheet.cell(row=count, column=7).value.lower() == 'high yielding variety':
                            self.seed_Details_part2[0] += 1
                            self.seed_Details_part2_counter += 1
                        elif sheet.cell(row=count, column=7).value.lower() == 'local':
                            self.seed_Details_part2[1] += 1
                            self.seed_Details_part2_counter += 1
                        elif sheet.cell(row=count, column=7).value.lower() == 'hybrid':
                            self.seed_Details_part2[2] += 1
                            self.seed_Details_part2_counter += 1

                        else:
                            print('Some problem in checking the seed type part 2')

                        if type(sheet.cell(row=count, column=24).value) == float or type(
                                sheet.cell(row=count, column=24).value) == int:
                            self.pestDamage += sheet.cell(row=count, column=24).value
                            self.pestDamage_counter += 1

                    if variety != sheet.cell(row=count, column=8).value and checker2 == True:
                        # print('reached here 11')

                        ratio = self.isIrrigatedType[0] / self.isIrrigatedType_counter
                        sheet2.cell(row=count2, column=8, value=ratio)
                        # print(ratio)
                        ratio = self.isIrrigatedType[1] / self.isIrrigatedType_counter
                        sheet2.cell(row=count2, column=9, value=ratio)
                        # print(ratio)
                        ratio = self.isIrrigatedType[2] / self.isIrrigatedType_counter
                        sheet2.cell(row=count2, column=10, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part1[0] / self.seed_Details_part1_counter
                        sheet2.cell(row=count2, column=11, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part1[1] / self.seed_Details_part1_counter
                        sheet2.cell(row=count2, column=12, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part2[0] / self.seed_Details_part2_counter
                        sheet2.cell(row=count2, column=13, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part2[1] / self.seed_Details_part2_counter
                        sheet2.cell(row=count2, column=14, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part2[2] / self.seed_Details_part2_counter
                        sheet2.cell(row=count2, column=15, value=ratio)
                        try:
                            ratio = self.pestDamage / self.pestDamage_counter
                        except ZeroDivisionError:
                            sheet2.cell(row=count2, column=16, value=0)
                        else:
                            sheet2.cell(row=count2, column=16, value=ratio)

                        ratio = 0
                        # initialising all to zero

                        self.isIrrigatedType = [0, 0, 0];
                        self.isIrrigatedType_counter = 0
                        self.seed_Details_part1 = [0, 0];
                        self.seed_Details_part1_counter = 0
                        self.seed_Details_part2 = [0, 0, 0];
                        self.seed_Details_part2_counter = 0
                        self.pestDamage_counter = 0;
                        self.pestDamage = 0
                        checker2 = False
                        break
                    if sheet.cell(row=count, column=8).value == None: break
            file2.save('Blocks/'+self.blockNames[jj]+'/ratios/ratio_normal_' + self.globalWorkBookName + '.xlsx')
            file2.close()
            file.close()
        self.reason_good()

    def reason_good(self):
        for jj in range(0, len(self.blockNames)):
            file = load_workbook('Blocks/'+self.blockNames[jj]+'/sorts/sortedFile_Good_' + self.globalWorkBookName + '_.xlsx')
            sheet = file['Sorted']
            self.isIrrigatedType = [0] * 3
            self.isIrrigatedType_counter = 0

            # irrigated, rainfed, unirrigated
            self.seed_Details_part1 = [0] * 2
            self.seed_Details_part1_counter = 0

            # conventional , sri
            self.seed_Details_part2 = [0] * 3
            self.seed_Details_part2_counter = 0

            # high yielding, local, hybrid
            count = 1

            # pests damage avergae calculation
            self.pestDamage = 0
            self.pestDamage_counter = 0

            file2 = load_workbook('Blocks/'+self.blockNames[jj]+'/ratios/ratio_good_' + self.globalWorkBookName + '.xlsx')
            sheet2 = file2['Sheet']
            count2 = 1
            sheet2.cell(row=1, column=8, value='irrigated ratio')
            sheet2.cell(row=1, column=9, value='rainfed ratio')
            sheet2.cell(row=1, column=10, value='unirrigated ratio')
            sheet2.cell(row=1, column=11, value='conventional ratio')
            sheet2.cell(row=1, column=12, value='sri ratio')
            sheet2.cell(row=1, column=13, value='high yielding ratio')
            sheet2.cell(row=1, column=14, value='local ratio')
            sheet2.cell(row=1, column=15, value='hybrid ratio')
            sheet2.cell(row=1, column=16, value='pest damage')

            while sheet2.cell(row=count2, column=4).value != None:
                # print('Entered sheet2 part')
                count2 += 1
                variety = sheet2.cell(row=count2, column=4).value
                if variety == None: break

                checker2 = False
                count = 2
                while sheet.cell(row=count, column=2).value != None:
                    count = count + 1
                    # print('sheet1 entered')
                    # irrigation type part below
                    if variety == sheet.cell(row=count, column=8).value:

                        checker2 = True
                        if sheet.cell(row=count, column=21).value.lower() == 'irrigated':
                            self.isIrrigatedType[0] += 1
                            self.isIrrigatedType_counter += 1
                        elif sheet.cell(row=count, column=21).value.lower() == 'rainfed':
                            self.isIrrigatedType[1] += 1
                            self.isIrrigatedType_counter += 1
                        elif sheet.cell(row=count, column=21).value.lower() == 'un-irrigated' or \
                                sheet.cell(row=count, column=21).value.lower() == 'un irrigated':
                            self.isIrrigatedType[2] += 1
                            self.isIrrigatedType_counter += 1
                        else:
                            print('Some problem in checking the irrigation type from the database')
                            print('this : ' + sheet.cell(row=count, column=21).value)

                        # seed details part1 below

                        if sheet.cell(row=count, column=6).value.lower() == 'conventional':
                            self.seed_Details_part1[0] += 1
                            self.seed_Details_part1_counter += 1
                        elif sheet.cell(row=count, column=6).value.lower() == 'sri':
                            self.seed_Details_part1[1] += 1
                            self.seed_Details_part1_counter += 1
                        else:
                            print('Some problem in checking the seed type part 1')

                        # seed details part2 below

                        if sheet.cell(row=count, column=7).value.lower() == 'high yielding variety':
                            self.seed_Details_part2[0] += 1
                            self.seed_Details_part2_counter += 1
                        elif sheet.cell(row=count, column=7).value.lower() == 'local':
                            self.seed_Details_part2[1] += 1
                            self.seed_Details_part2_counter += 1
                        elif sheet.cell(row=count, column=7).value.lower() == 'hybrid':
                            self.seed_Details_part2[2] += 1
                            self.seed_Details_part2_counter += 1

                        else:
                            print('Some problem in checking the seed type part 2')

                        if type(sheet.cell(row=count, column=24).value) == float or type(
                                sheet.cell(row=count, column=24).value) == int:
                            self.pestDamage += sheet.cell(row=count, column=24).value
                            self.pestDamage_counter += 1

                    if variety != sheet.cell(row=count, column=8).value and checker2 == True:
                        # print('reached here 11')

                        ratio = self.isIrrigatedType[0] / self.isIrrigatedType_counter
                        sheet2.cell(row=count2, column=8, value=ratio)
                        # print(ratio)
                        ratio = self.isIrrigatedType[1] / self.isIrrigatedType_counter
                        sheet2.cell(row=count2, column=9, value=ratio)
                        # print(ratio)
                        ratio = self.isIrrigatedType[2] / self.isIrrigatedType_counter
                        sheet2.cell(row=count2, column=10, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part1[0] / self.seed_Details_part1_counter
                        sheet2.cell(row=count2, column=11, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part1[1] / self.seed_Details_part1_counter
                        sheet2.cell(row=count2, column=12, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part2[0] / self.seed_Details_part2_counter
                        sheet2.cell(row=count2, column=13, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part2[1] / self.seed_Details_part2_counter
                        sheet2.cell(row=count2, column=14, value=ratio)
                        # print(ratio)
                        ratio = self.seed_Details_part2[2] / self.seed_Details_part2_counter
                        sheet2.cell(row=count2, column=15, value=ratio)
                        try:
                            ratio = self.pestDamage / self.pestDamage_counter
                        except ZeroDivisionError:
                            sheet2.cell(row=count2, column=16, value=0)
                        else:
                            sheet2.cell(row=count2, column=16, value=ratio)

                        ratio = 0
                        # initialising all to zero

                        self.isIrrigatedType = [0, 0, 0];
                        self.isIrrigatedType_counter = 0
                        self.seed_Details_part1 = [0, 0];
                        self.seed_Details_part1_counter = 0
                        self.seed_Details_part2 = [0, 0, 0];
                        self.seed_Details_part2_counter = 0
                        self.pestDamage_counter = 0;
                        self.pestDamage = 0
                        checker2 = False
                        break
                    if sheet.cell(row=count, column=8).value == None: break
            file2.save('Blocks/'+self.blockNames[jj]+'/ratios/ratio_good_' + self.globalWorkBookName + '.xlsx')
            file2.close()
            file.close()





if __name__  ==  '__main__' :
    obj = Block_Division()
    obj.individualAllotment()
    obj.process_General()
    obj.xx()
    obj.avg_Good()
    obj.bad()
    obj.opening()



