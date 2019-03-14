from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
from math import floor
import numpy as np
import csv

class Predictor():
    def __init__(self):
        #referal_book = input('Enter the referal book name (without any extension) : ')
        referal_book = 'C3'
        self.ratio_bad3 = load_workbook('ratios/ratio_bad_' + referal_book + '.xlsx')
        self.ratio_normal3 = load_workbook('ratios/ratio_normal_' + referal_book + '.xlsx')
        self.ratio_good3 = load_workbook('ratios/ratio_good_' + referal_book + '.xlsx')
        self.ratio_good2 = self.ratio_good3['Sheet']
        self.ratio_normal2 = self.ratio_normal3['Sheet']
        self.ratio_bad2 = self.ratio_bad3['Sheet']
        self.ratio_bad = pd.read_excel('ratios/ratio_bad_' + referal_book + '.xlsx')
        self.ratio_normal = pd.read_excel('ratios/ratio_normal_' + referal_book + '.xlsx')
        self.ratio_good = pd.read_excel('ratios/ratio_good_' + referal_book + '.xlsx')

        self.blockname = ''
        self.vareityname = ''
        self.systemCultivation = ''
        self.isIrrigated = ''
        self.waterSource = ''
        self.pestDamage = ''
        self.operationSize = 0  # in hectare
        self.cultivatedSize = 0  # in hectare
        self.ask_from_Excel()

    def ask_from_Excel(self):
        print('Make sure, your csv file has the following heading as the columns headings:')
        print('variety name\nsystem of cultivation\nis irrigated\nyielding type\npest damage\nseeds per hectare\noperation size\ncultivation size')

        self.name = input('Csv file name to be Predicted ( without extension ) : ')
        self.ca = self.name
        self.name = self.name + '.csv'
        file = pd.read_csv(self.name)
        self.vareityname2 = file['variety name']
        self.systemCultivation2 = file['system of cultivation']
        self.isIrrigated2 = file['is irrigated']
        self.yieldingType2 = file['yielding type']
        self.pestDamage2 = file['pest damage']
        self.seedPerHect2 = file['seeds per hectare']
        self.operationSize2 = file['operation size']
        self.cultivatedSize2 = file['cultivation size']

        ## new file for storage

        self.file22 = open(self.ca + '_predicted_.csv', 'w')
        self.writer = csv.writer(self.file22, dialect='excel')
        w1 = ['variety name', 'system of cultivation', 'is irrigated', 'yielding type', 'pest damage', 'seeds per hectare', 'operation size'
              , 'cultivation size' , 'expected green wt in CCE','expected dry wt in CCE', 'expected normal yield in kilograms','result / remarks'
              ]
        self.writer.writerow(w1)
        self.pre_normal_storage = []
        self.pre_greenwt_storage = []
        self.transporter()


    '''
    def EnterValues(self):
        print('Please input the required values')

        self.blockname = input('Block Name : ')
        self.vareityname = input('Variety Name : ').lower()
        print(self.vareityname)
        #self.vareityname = 'puja'
        #self.systemCultivation = input('System of Cultivation : ').lower
        self.systemCultivation = 'conventional'
        self.isIrrigated = input('is irrigated (irrigated/rainfed/unirrigated): ').lower()
        #self.isIrrigated = 'rainfed'
        self.yieldingType = input('yielding type (local/hybrid/high yielding)' ).lower()
        #self.yieldingType = 'local'
        # self.waterSource = input('Water Source : ')
        self.pestDamage = int(input('Pest Damage in Integers : '))
        #self.pestDamage = 0
        self.seedPerHect = float(input('Seeds Sown Per Hectare : '))
        #self.seedPerHect = 5
        self.operationSize = float(input('Operational Size holding (in hectares) : '  ))
        #self.operationSize = 3.0
        self.cultivatedSize = float(input('Operational Size in Cultivation (in hectares) : '  ))
        #self.cultivatedSize = 1.5
        self.calculatingWeightage()
    '''

    def transporter(self):
        for i in range(len(self.vareityname2)):
            self.vareityname = self.vareityname2[i]
            self.systemCultivation = self.systemCultivation2[i]
            self.isIrrigated = self.isIrrigated2[i]
            self.yieldingType = self.yieldingType2[i]
            self.pestDamage = self.pestDamage2[i]
            self.seedPerHect = self.seedPerHect2[i]
            self.operationSize = self.operationSize2[i]
            self.cultivatedSize = self.cultivatedSize2[i]
            self.presentRow = i + 2
            self.calculatingWeightage()
        self.visualisation()

    def calculatingWeightage(self):

        self.varietyRow_number = [0] * 3
        i = 1
        check = False
        while True:
            i += 1
            if self.ratio_bad2.cell(row=i, column=4).value == None:
                break

            v = self.ratio_bad2.cell(row=i, column=4).value
            if self.vareityname == v:
                self.varietyRow_number[0] = i
                check = True
                break
        if check == False:
            print('Variety not found in the bads file')
        i = 1
        check = False
        while True:
            i += 1
            if self.ratio_normal2.cell(row=i, column=4).value == None:
                break

            v = self.ratio_normal2.cell(row=i, column=4).value
            if self.vareityname == v:
                self.varietyRow_number[1] = i
                check = True
                break
        if check == False:
            print('Variety not found in the normals file')
        i = 1
        check = False
        while True:
            i += 1
            if self.ratio_good2.cell(row=i, column=4).value == None:
                break

            v = self.ratio_good2.cell(row=i, column=4).value
            if self.vareityname == v:
                self.varietyRow_number[2] = i
                check = True
                break
        if check == False:
            print('Variety not found in the goods file')

        i = 1

        ### *********** deciding Category ***************

        # bad file part
        sum1 = 0
        if self.isIrrigated == 'irrigated':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=8).value
        elif self.isIrrigated == 'rainfed':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=9).value
        elif self.isIrrigated == 'unirrigated':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=10).value
        else:
            print('Wrong input in irrigation part')
            exit(0)

        if self.systemCultivation == 'conventional':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=11).value
        elif self.systemCultivation == 'sri':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=12).value
        else:
            print('wrong input in method of cultivation part')
            exit(0)

        if self.yieldingType == 'high yielding':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=13).value
        elif self.yieldingType == 'local':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=14).value
        elif self.yieldingType == 'hybrid':
            sum1 += self.ratio_bad2.cell(row=self.varietyRow_number[0], column=15).value
        else:
            print('wrong input in variety part')
            exit(0)

        # normal file part

        sum2 = 0
        if self.isIrrigated == 'irrigated':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=8).value
        elif self.isIrrigated == 'rainfed':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=9).value
        elif self.isIrrigated == 'unirrigated':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=10).value
        else:
            print('Wrong input in irrigation part')
            exit(0)

        if self.systemCultivation == 'conventional':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=11).value
        elif self.systemCultivation == 'sri':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=12).value
        else:
            print('wrong input in method of cultivation part')
            exit(0)

        if self.yieldingType == 'high yielding':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=13).value
        elif self.yieldingType == 'local':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=14).value
        elif self.yieldingType == 'hybrid':
            sum2 += self.ratio_normal2.cell(row=self.varietyRow_number[1], column=15).value
        else:
            print('wrong input in variety part')
            exit(0)

        # good file part

        sum3 = 0
        if self.isIrrigated == 'irrigated':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=8).value
        elif self.isIrrigated == 'rainfed':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=9).value
        elif self.isIrrigated == 'unirrigated':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=10).value
        else:
            print('Wrong input in irrigation part')
            exit(0)

        if self.systemCultivation == 'conventional':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=11).value
        elif self.systemCultivation == 'sri':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=12).value
        else:
            print('wrong input in method of cultivation part')
            exit(0)

        if self.yieldingType == 'high yielding':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=13).value
        elif self.yieldingType == 'local':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=14).value
        elif self.yieldingType == 'hybrid':
            sum3 += self.ratio_good2.cell(row=self.varietyRow_number[2], column=15).value
        else:
            print('wrong input in variety part')
            exit(0)

        ### *********** pest importance calculation ****************

        val = self.pestDamage
        valpest_bad = self.ratio_bad2.cell(row=self.varietyRow_number[0], column=16).value
        valpest_normal = self.ratio_normal2.cell(row=self.varietyRow_number[1], column=16).value
        valpest_good = self.ratio_good2.cell(row=self.varietyRow_number[2], column=16).value

        d_bad = abs(val - valpest_bad)
        d_normal = abs(val - valpest_normal)
        d_good = abs(val - valpest_good)
        self.result = ''

        if d_bad < d_normal and d_bad < d_good:
            sum1 += 2
            self.result = 'bad'
        elif d_normal < d_bad and d_normal < d_good:
            sum2 += 2
            self.result = 'normal'
        elif d_good < d_bad and d_good < d_normal:
            sum3 += 2
            self.result = 'good'
        else:
            print('Cannot decide the pestDamage Factor. Please Report the Bug team.')

        self.rowCarry = 0

        print('Sum 1 :' + str(sum1) + '\nSum 2 : ' + str(sum2) + '\nSum 3 : ' + str(sum3))

        if self.result == 'bad':
            self.predictingBad(self.varietyRow_number[0])
        elif self.result == 'normal':
            self.predictingNormal(self.varietyRow_number[1])
        elif self.result == 'good':
            self.predictingGood(self.varietyRow_number[2])
        else:
            print('self.result Calculation Failed.! Please report the Bug Team about this.')

        self.array_com11 = []
        self.nominalValues()

    def predictingBad(self, rowNumber):
        print('Got the BAD part')
        self.pre_greenWtProduced = self.ratio_bad2.cell(row=rowNumber,
                                                        column=1).value * self.cultivatedSize * self.seedPerHect
        self.pre_dryWt = self.ratio_bad2.cell(row=rowNumber, column=2).value * self.cultivatedSize * self.seedPerHect
        self.pre_normalYieldinKilo = self.ratio_bad2.cell(row=rowNumber,
                                                          column=3).value * self.cultivatedSize * self.seedPerHect
        self.rowCarry = rowNumber
        print('Expected Values :')
        print(
            'Green Weight Produced would be ' + str(self.pre_greenWtProduced) + '\nDry Weight Produced would be ' + str(
                self.pre_dryWt) \
            + '\nNormal Yield in Kilograms would be ' + str(self.pre_normalYieldinKilo)
            )
        w = [
            self.vareityname, self.systemCultivation, self.isIrrigated, self.yieldingType, self.pestDamage, self.seedPerHect,
            self.operationSize, self.cultivatedSize,self.pre_greenWtProduced,self.pre_dryWt,self.pre_normalYieldinKilo,self.result
        ]
        print(w)
        self.pre_normal_storage.append(self.pre_normalYieldinKilo)
        self.pre_greenwt_storage.append(self.pre_greenWtProduced)
        self.writer.writerow(w)


    def predictingNormal(self, rowNumber):
        print('Got the NORMAL part')
        self.pre_greenWtProduced = self.ratio_normal2.cell(row=rowNumber,
                                                           column=1).value * self.cultivatedSize * self.seedPerHect
        self.pre_dryWt = self.ratio_normal2.cell(row=rowNumber, column=2).value * self.cultivatedSize * self.seedPerHect
        self.pre_normalYieldinKilo = self.ratio_normal2.cell(row=rowNumber,
                                                             column=3).value * self.cultivatedSize * self.seedPerHect
        self.rowCarry = rowNumber
        print('Expected Values :')
        print(
            'Green Weight Produced would be ' + str(self.pre_greenWtProduced) + '\nDry Weight Produced would be ' + str(
                self.pre_dryWt) \
            + '\nNormal Yield in Kilograms would be ' + str(self.pre_normalYieldinKilo)
            )
        w = [
            self.vareityname, self.systemCultivation, self.isIrrigated, self.yieldingType, self.pestDamage,
            self.seedPerHect,
            self.operationSize, self.cultivatedSize,self.pre_greenWtProduced,self.pre_dryWt,self.pre_normalYieldinKilo, self.result
        ]
        print(w)
        self.pre_normal_storage.append(self.pre_normalYieldinKilo)
        self.pre_greenwt_storage.append(self.pre_greenWtProduced)
        self.writer.writerow(w)

    def predictingGood(self, rowNumber):
        print('Got the GOOD part')
        self.pre_greenWtProduced = self.ratio_good2.cell(row=rowNumber,
                                                         column=1).value * self.cultivatedSize * self.seedPerHect
        self.pre_dryWt = self.ratio_good2.cell(row=rowNumber, column=2).value * self.cultivatedSize * self.seedPerHect
        self.pre_normalYieldinKilo = self.ratio_good2.cell(row=rowNumber,
                                                           column=3).value * self.cultivatedSize * self.seedPerHect
        self.rowCarry = rowNumber
        print('Expected Values :')
        print(
            'Green Weight Produced would be ' + str(self.pre_greenWtProduced) + '\nDry Weight Produced would be ' + str(
                self.pre_dryWt) \
            + '\nNormal Yield in Kilograms would be ' + str(self.pre_normalYieldinKilo)
            )
        w = [
            self.vareityname, self.systemCultivation, self.isIrrigated, self.yieldingType, self.pestDamage,
            self.seedPerHect,
            self.operationSize, self.cultivatedSize,self.pre_greenWtProduced,self.pre_dryWt,self.pre_normalYieldinKilo, self.result
        ]
        print(w)
        self.pre_normal_storage.append(self.pre_normalYieldinKilo)
        self.pre_greenwt_storage.append(self.pre_greenWtProduced)
        self.writer.writerow(w)

    def visualisation(self):
        x = self.vareityname2
        y = self.pre_normal_storage
        z = np.array(self.pre_greenwt_storage)
        z = z / 100
        plt.scatter(x,y,z)
        plt.xlabel('Variety name')
        plt.ylabel('Normal yield in kilograms')
        plt.title('Variety Name from csv given vs Predicted Normal yield vs Green Wt Predicted of the given csv(in bubbles) ')
        plt.show()

    def nominalValues(self):
        self.d_perHectareSeed = 50
        self.d_manure = 150
        #self.perHectGraph()

    '''

    ## ***************** VISUALISATION ********************

    def pre_greenWt_Engine(self, fuel):
        #print(self.rowCarry)
        print(self.pre_greenWtProduced)
        xx = (self.pre_greenWtProduced/self.seedPerHect) * fuel
        return xx

    def pre_dryWt_Engine(self, fuel):
        yy = (self.pre_dryWt/self.seedPerHect) * fuel
        return yy

    def pre_normalYield_Engine(self, fuel):
        zz = (self.pre_normalYieldinKilo/self.seedPerHect) * fuel
        return zz

    def perHectGraph(self):
        temporary = self.seedPerHect
        self.seedPerHect = floor(self.seedPerHect)
        diff = abs(self.d_perHectareSeed - self.seedPerHect)

        array_com1 = []
        if diff == 0:
            for i in range(1, 101):
                array_com1.append(i)
                self.array_com11.append(i)
            # self.array_com11 = array_com1
            print('Temp is :')
            print(self.array_com11)
            count = 0
            b = 50
            for i in self.array_com11:
                if i >= 50:

                    array_com1[count] = b
                    b = b - 1
                else:
                    array_com1[count] = i
                count += 1

        else:
            if diff >= 30:
                for i in range(diff - 10, 101):
                    array_com1.append(i)
                    self.array_com11.append(i)
                # self.array_com11 = array_com1
                print('Temp is :')
                print(self.array_com11)
                count = 0
                b = 50
                for i in self.array_com11:
                    if i >= 50:

                        array_com1[count] = b
                        b = b - 1
                    else:
                        array_com1[count] = i
                    count += 1

            else:
                for i in range(diff + 10, 101):
                    array_com1.append(i)
                    self.array_com11.append(i)
                # self.array_com11 = array_com1
                print('Temp is :')
                print(self.array_com11)
                count = 0
                b = 50
                for i in self.array_com11:
                    if i >= 50:

                        array_com1[count] = b
                        b = b - 1
                    else:
                        array_com1[count] = i
                    count += 1

        array_com2 = []
        array_com_dry = []
        array_com_normal = []

        for i in array_com1:
            x = self.pre_greenWt_Engine(i)
            array_com2.append(x)
            y = self.pre_dryWt_Engine(i)
            array_com_dry.append(y)
            z = self.pre_normalYield_Engine(i)
            array_com_normal.append(z)


        # array_com11 = self.array_com11

        print(self.array_com11)
        print(array_com2)
        print(array_com1)
        print('dry and then normal')
        print(array_com_dry)
        print(array_com_normal)




        ## ********* making Graph **************
    
        plt.plot(self.array_com11, array_com2)
        plt.xlabel('Seeds in KiloGrams per Hectare ')
        plt.ylabel('Expected Green Wt in CCE in Kilograms')
        plt.title('Prediction of Green Weight in CCE produce based on Variation of Seeds per Hectare ')
        plt.text(50,0,'Your per Hectare seed was ' + str(temporary) +' Kilograms')
        plt.show()
        
        fig = plt.figure()
        f1 = fig.add_subplot(111)
        f1.fill(self.array_com11, array_com2)
        f1.text(0,0,'EXPECTED GREEN WEIGHT PRODUCED vs SEED PER HECTARE VARIANCE')
        fig2 = plt.figure()
        f2 = fig2.add_subplot(111)
        f2.fill(self.array_com11, array_com_dry)
        f2.text(0,0,'EXPECTED DRY WEIGHT PRODUCED CCE vs SEED PER HECTARE VARIANCE')
        fig3 = plt.figure()
        f3 = fig3.add_subplot(111)
        nn = np.array(array_com_normal)
        nn = nn /1000
        f3.fill(self.array_com11, nn)
        f3.text(0,0,'EXPECTED NORMAL YIELDING in thousands kilos vs SEED PER HECTARE VARIANCE')
        fig4 = plt.figure()
        f4 = fig4.add_subplot(111)
        f4.plot(array_com2, nn)
        f4.text(0,0,'GREEN WEIGHT PRODUCED vs DRY WEIGHT PRODUCED ')
        plt.show()

        '''




if __name__ == '__main__':
    obj = Predictor()