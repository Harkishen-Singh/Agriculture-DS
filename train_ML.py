from openpyxl import  load_workbook
import csv
import matplotlib.pyplot as plt
import numpy as np

class Training_ML():

    def __init__(self):
        self.workbook_name = input('Enter SpreadSheet name (extension .xlsx): ')
        self.workbook_name = './'+self.workbook_name+'.xlsx'
        self.wb = load_workbook(self.workbook_name)
        self.ws = self.wb['Sheet1']
        self.blockName = []
        self.villageName = []
        self.areaHolding = []
        self.areaCultivation = []
        self.sysCultivation = []
        self.cropVariety = []
        self.varietyName = []
        self.sourceSeed = []
        self.seedpHectare = []
        self.manureFYMq = []
        self.chemFertiQ = []
        self.dateTransplanting = []
        self.dateHarvest = []
        self.greenWtCCE = []
        self.dryWtCCE = []
        self.moisturePercent = []
        self.normalYldQuin = [] # while storing, convert into kilo and delete this variable
        self.normalYldKilo = []
        self.proQuantityUser = []
        self.proQualityUser = []
        self.isIrrigated = []
        self.waterSource = []
        self.weatherCond = []
        self.pest = []
        self.stress = []
        self.headingsSheet = []
        self.headingIndex = ['']*25
        self.dateCutting = []
        # knowing total number of rows
        self.numberOfRows = 0
        self.numberOfColumns = 0
        for i in range(2,10000000):
            if self.ws.cell(row=i, column=1).value == None:
                self.numberOfRows = i-1
                break
        for i in range(2,10000000):
            if self.ws.cell(row=1, column=i).value == None:
                self.numberOfColumns = i-1
                break
        self.scannerHeadingNames()



    def scannerHeadingNames(self):
        self.nameSheet = self.ws
        self.correctInputFormat = [False] * 25
        for i in range(1, self.numberOfColumns+1):
            self.headingsSheet.append(self.ws.cell(row=1, column=i).value)
        for heads in range(1, self.numberOfColumns+1):
            if self.ws.cell(row=1, column=heads).value == 'Block_Name': # 0
                self.headingIndex[0]=heads
                self.correctInputFormat[0] = True
            if self.ws.cell(row=1, column=heads).value == 'Village_Name': # 1
                self.headingIndex[1]= heads
                self.correctInputFormat[1] = True
            if self.ws.cell(row=1, column=heads).value == 'Operational Size of the holding of farmer': # 2
                self.headingIndex[2] = heads
                self.correctInputFormat[2] = True
            if self.ws.cell(row=1, column=heads).value == 'Total area under crop(hectare) in respect of cultivator': # 3
                self.headingIndex[3] = heads
                self.correctInputFormat[3] = True
            if self.ws.cell(row=1, column=heads).value == 'System of Cultivation': # 4
                self.headingIndex[4] = heads
                self.correctInputFormat[4] = True
            if self.ws.cell(row=1, column=heads).value == 'Crop Variety Type': # 5
                self.headingIndex[5] = heads
                self.correctInputFormat[5] = True
            if self.ws.cell(row=1, column=heads).value == 'Variety Name':# 6
                self.headingIndex[6] = heads
                self.correctInputFormat[6] = True
            if self.ws.cell(row=1, column=heads).value == 'Sources of Seed':# 7
                self.headingIndex[7] = heads
                self.correctInputFormat[7] = True
            if self.ws.cell(row=1, column=heads).value == 'Seed Used per Hectare':# 8
                self.headingIndex[8] = heads
                self.correctInputFormat[8] = True
            if self.ws.cell(row=1, column=heads).value == 'Quantity of Manure/FYM used in(per hectare)':# 9
                self.headingIndex[9] = heads
                self.correctInputFormat[9] = True
            if self.ws.cell(row=1, column=heads).value == 'Quantity of Chemical Fertilizer (In hectare)':# 10
                self.headingIndex[10] = heads
                self.correctInputFormat[10] = True
            if self.ws.cell(row=1, column=heads).value == 'Time of showing or Transplanting': # 11
                self.headingIndex[11] = heads
                self.correctInputFormat[11] = True
            if self.ws.cell(row=1, column=heads).value == 'Date of likely harvest':# 12
                self.headingIndex[12] = heads
                self.correctInputFormat[12] = True
            if self.ws.cell(row=1, column=heads).value == 'Green Weight of the produce obtained in CCE':# 13
                self.headingIndex[13] = heads
                self.correctInputFormat[13] = True
            if self.ws.cell(row=1, column=heads).value == "Dry weight of the produce in CCE's":# 14
                self.headingIndex[14] = heads
                self.correctInputFormat[14] = True
            if self.ws.cell(row=1, column=heads).value == "Moisture percentage in the produce obtained in CCE's":# 15
                self.headingIndex[15] = heads
                self.correctInputFormat[15] = True
            if self.ws.cell(row=1, column=heads).value == "Normal Average yeild (in Quintals/Hectare)":# 16
                self.headingIndex[16] = heads
                self.correctInputFormat[16] = True
            if self.ws.cell(row=1, column=heads).value == "Production obtained through CCE's":# 17
                self.headingIndex[17] = heads
                self.correctInputFormat[17] = True
            if self.ws.cell(row=1, column=heads).value == "Remarks about production observed":# 18
                self.headingIndex[18] = heads
                self.correctInputFormat[18] = True
            if self.ws.cell(row=1, column=heads).value == "Is Field Irrigated":# 19
                self.headingIndex[19] = heads
                self.correctInputFormat[19] = True
            if self.ws.cell(row=1, column=heads).value == "Water Source":# 20
                self.headingIndex[20] = heads
                self.correctInputFormat[20] = True
            if self.ws.cell(row=1, column=heads).value == "Weather Condition during Crop Season":# 21
                self.headingIndex[21] = heads
                self.correctInputFormat[21] = True
            if self.ws.cell(row=1, column=heads).value == "Extent of damage by pests or any disease":# 22
                self.headingIndex[22] = heads
                self.correctInputFormat[22] = True
            if self.ws.cell(row=1, column=heads).value == "Any Stress":# 23
                self.headingIndex[23] = heads
                self.correctInputFormat[23] = True
            if self.ws.cell(row=1, column=heads).value == "Date Of Cutting":# 24
                self.headingIndex[24] = heads
                self.correctInputFormat[24] = True

        # error in resent column finder
        count2 =0
        for i in range(0, 25):
            if self.correctInputFormat[i] == False:
                if count2 == 0:
                    count2 = count2 + 1
                    print('Error in the format of excel file')
                print('Absence of ' +  self.ws.cell(row=1, column= i+1).value + ' column')

        self.assignment()

    def assignment(self): # to assign the values to the arrays declared above
        for i in range(2, self.numberOfRows+1):
            self.blockName.append(self.ws.cell(row=i, column=self.headingIndex[0]).value)
            self.villageName.append(self.ws.cell(row=i, column=self.headingIndex[1]).value)
            self.areaHolding.append(self.ws.cell(row=i, column=self.headingIndex[2]).value)
            self.areaCultivation.append(self.ws.cell(row=i, column=self.headingIndex[3]).value)
            self.sysCultivation.append(self.ws.cell(row=i, column=self.headingIndex[4]).value)
            self.cropVariety.append(self.ws.cell(row=i, column=self.headingIndex[5]).value)
            self.varietyName.append(self.ws.cell(row=i, column=self.headingIndex[6]).value)
            self.sourceSeed.append(self.ws.cell(row=i, column=self.headingIndex[7]).value)
            self.seedpHectare.append(self.ws.cell(row=i, column=self.headingIndex[8]).value)
            self.manureFYMq.append(self.ws.cell(row=i, column=self.headingIndex[9]).value)
            self.chemFertiQ.append(self.ws.cell(row=i, column=self.headingIndex[10]).value)
            self.dateTransplanting.append(self.ws.cell(row=i, column=self.headingIndex[11]).value)
            self.dateHarvest.append(self.ws.cell(row=i, column=self.headingIndex[12]).value)
            self.greenWtCCE.append(self.ws.cell(row=i, column=self.headingIndex[13]).value)
            self.dryWtCCE.append(self.ws.cell(row=i, column=self.headingIndex[14]).value)
            self.moisturePercent.append(self.ws.cell(row=i, column=self.headingIndex[15]).value)
            self.normalYldKilo.append(self.ws.cell(row=i, column=self.headingIndex[16]).value*100)
            self.proQuantityUser.append(self.ws.cell(row=i, column=self.headingIndex[17]).value)
            self.proQualityUser.append(self.ws.cell(row=i, column=self.headingIndex[18]).value)
            self.isIrrigated.append(self.ws.cell(row=i, column=self.headingIndex[19]).value)
            self.waterSource.append(self.ws.cell(row=i, column=self.headingIndex[20]).value)
            self.weatherCond.append(self.ws.cell(row=i, column=self.headingIndex[21]).value)
            self.pest.append(self.ws.cell(row=i, column=self.headingIndex[22]).value)
            self.stress.append(self.ws.cell(row=i, column=self.headingIndex[23]).value)
            self.dateCutting.append(self.ws.cell(row=i, column=self.headingIndex[24]).value)

        self.numpyConversion()

    def numpyConversion(self):
        # this assigns the numpy array with the required values for the processing
        self.np_main = np.array((
            self.seedpHectare,
            self.manureFYMq,
            self.chemFertiQ,
            self.greenWtCCE,
            self.dryWtCCE,
            self.normalYldKilo,
            self.moisturePercent,
            self.pest
        )
        )





